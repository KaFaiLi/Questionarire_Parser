"""
PDF (vision-based) Questionnaire Parser
=======================================
Parses questionnaires by converting the xlsx to PDF (via Excel + pywin32),
rendering each page to a PNG, and asking an Azure OpenAI vision model (via
langchain) to extract the question_id / question / answer triples. Read-only /
locked PDFs are fine — pypdfium2 renders them without modification.

The xlsx→pdf step only keeps columns **B:AN** on every sheet (the print area is
restricted before export) so the irrelevant left margin (col A) and any noise
to the right of AN never reach the LLM.

Why a sliding window?
---------------------
Long questionnaires get split across many pages, and a single question (or its
answer) can land at the bottom of one page and continue at the top of the next.
To make sure every Q/A is seen in full at least once, we send the LLM
overlapping pairs of consecutive pages: [p1, p2], [p2, p3], [p3, p4], ...
Each Q/A then appears in at least one window where it is not cut. After all
windows are processed, results are deduped by question_id, preferring the most
complete (longest non-empty answer / longest question text) entry.

Inputs
------
A .xlsx file or a directory of .xlsx files. (A .pdf may also be passed
directly, in which case the xlsx→pdf step is skipped.)

Outputs
-------
For each input file, in ``--output-dir``:
    <name>.json       – {file_name, questions:[{question_id, question,
                         sub_questions:[{option_label, prompt, selection,
                         answer, answer_source}], source_pages}]}. Each
                         answerable part of a question (multi-part 'a)/b)',
                         a YES/NO branch, a single free-text box) is its own
                         sub-question so selections and answers never have to
                         be collapsed into one field.
    <name>.pdf        – the intermediate PDF (kept for inspection / debugging)
And a combined:
    <stem>.xlsx       – one row per sub-question across all inputs

Environment
-----------
AZURE_OPENAI_API_KEY        – API key
AZURE_OPENAI_ENDPOINT       – e.g. https://my-resource.openai.azure.com/
AZURE_OPENAI_DEPLOYMENT     – chat deployment name (must be a vision model,
                              e.g. gpt-4o / gpt-4o-mini / gpt-4.1)
AZURE_OPENAI_API_VERSION    – e.g. 2024-10-21 (default if unset)

Dependencies (add to pyproject.toml or install separately):
    pip install pypdfium2 langchain langchain-openai openpyxl pillow pywin32
The xlsx→pdf step uses Excel COM automation via pywin32, so it requires
**Windows with Microsoft Excel installed**. PDF rendering and the LLM call
both work cross-platform.

Usage:
    python pdf_parse.py <file_or_dir> [--output-dir ./output] \
        [--dpi 200] [--window 2] [--stride 1] [--keep-pngs]
"""

from __future__ import annotations

import argparse
import base64
import io
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, TypeVar

import openpyxl
import pypdfium2 as pdfium
from PIL import Image
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_core.runnables import Runnable
from langchain_openai import AzureChatOpenAI
from pydantic import BaseModel, Field

# Helpers reused from the cell-based parser. The reconcile pass needs the same
# notion of "answer cell" (grey/yellow fill) and question-id pattern.
from parse import (
    ANSWER_FILLS,
    CONTENT_END_COL,
    CONTENT_START_COL,
    Q_PATTERN,
    build_merged_lookup,
    classify_fill,
    effective_value,
)

DEFAULT_API_VERSION = "2024-10-21"

# xlsx→pdf print-area limits (inclusive). Columns outside this range are
# excluded from every sheet before the PDF is exported.
PRINT_COL_START = "B"
PRINT_COL_END   = "AN"

# A column safely past PRINT_COL_END used as a measurement surface for the
# merged-cell row-height AutoFit pass. Nothing written here can reach the PDF.
SCRATCH_COL_LETTER = "AZ"

# Excel constant for ExportAsFixedFormat. (xlTypePDF = 0.)
XL_TYPE_PDF = 0

# LLM retry policy. We don't use langchain's built-in retry (`max_retries`)
# because it only retries certain transport errors — we want to retry on
# anything (404s, parsing/validation errors, rate limits, transient network
# failures) with explicit visibility in logs.
DEFAULT_MAX_ATTEMPTS = 4
DEFAULT_RETRY_BASE_DELAY = 2.0  # seconds; doubles each attempt

# ─── Data model the LLM must return ─────────────────────────────────────────
class SubQuestion(BaseModel):
    """One answerable part of a question.

    A single question id often bundles several of these: multi-part questions
    ('a) …', 'b) …'), branch/option rows (e.g. 'Settling distributor' vs
    'Non-settling distributor'), or just a single free-text box. Splitting them
    out means the LLM never has to pick one sub-part or merge several selections
    and answers into a single string.
    """
    option_label: str = Field(
        default="",
        description=(
            "The choice/branch this sub-question belongs to, exactly as printed "
            "(e.g. 'Settling distributor', 'Non-settling distributor'). Empty "
            "string if the sub-question is not under a labelled option."
        ),
    )
    prompt: str = Field(
        default="",
        description=(
            "The text of this specific sub-question / follow-up prompt (e.g. "
            "'Who is the custodian?' or 'b) Does each product go through the "
            "same approval process?'). For a plain question with a single "
            "answer box and no distinct sub-parts, leave this empty — the "
            "stem `question` already holds the text."
        ),
    )
    selection: str = Field(
        default="",
        description=(
            "The respondent's selected control for this sub-question: a side "
            "dropdown value (e.g. 'YES' / 'NO'), a ticked checkbox/radio label, "
            "or a highlighted choice. Empty string if there is no selectable "
            "control or none was chosen. Never put free text here."
        ),
    )
    answer: str = Field(
        default="",
        description=(
            "The respondent's free-text answer written in this sub-question's "
            "answer box. Empty string if the box is blank. Keep this separate "
            "from `selection`, and do not copy the prompt into it."
        ),
    )

class QA(BaseModel):
    question_id: str = Field(
        description=(
            "The question identifier exactly as printed (e.g. 'Q1', 'Q12', "
            "'1.a', '3.2'). If the page shows no explicit id but the row is "
            "clearly a question, use the visible numbering or a stable label."
        )
    )
    question: str = Field(
        description=(
            "The main/stem question text exactly as printed next to the "
            "question id (e.g. 'In which capacity will your firm act as "
            "distributor?'). Do NOT fold sub-prompts or option labels into "
            "this field — those belong in `sub_questions`."
        )
    )
    sub_questions: list[SubQuestion] = Field(
        default_factory=list,
        description=(
            "Every answerable part of this question, in the order it appears: "
            "one entry per option/branch row, per multi-part sub-question "
            "('a)…','b)…'), or — for a plain single-answer question — a single "
            "entry carrying just the answer."
        ),
    )

class QAList(BaseModel):
    questions: list[QA] = Field(default_factory=list)

# ─── xlsx → pdf ─────────────────────────────────────────────────────────────
def _sheet_last_row(ws) -> int:
    """Last row that has any content anywhere on the sheet, via Excel's
    UsedRange. Falls back to 1 if the sheet is empty."""
    used = ws.UsedRange
    if used is None:
        return 1
    return max(1, used.Row + used.Rows.Count - 1)

def _is_blank_value(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())

def _autofit_merged_rows(ws, last_row: int) -> int:
    """Grow row heights so each merged wrap-text cell fits its full text.

    Excel does NOT auto-fit row height for merged cells (a well-known
    limitation), so answer boxes with long wrapped text get visually clipped
    in the exported PDF. We measure the required height for each merged
    wrap-text cell using a scratch column past the export band, and grow
    the anchor row accordingly. Returns the number of rows adjusted.
    """
    scratch_col = ws.Columns(SCRATCH_COL_LETTER)
    scratch_col_index = scratch_col.Column
    original_width = scratch_col.ColumnWidth
    # Pick a measurement row far past any real content so the AutoFit on
    # that row only sees our scratch cell, not real cells we'd grow by
    # accident.
    measurement_row = last_row + 10
    seen: set[str] = set()
    adjusted = 0

    try:
        for cell in ws.UsedRange:
            area = cell.MergeArea
            if area.Cells.Count == 1:
                continue
            addr = area.Address
            if addr in seen:
                continue
            seen.add(addr)

            anchor = area.Cells(1, 1)
            if not anchor.WrapText:
                continue
            if _is_blank_value(anchor.Value):
                continue

            first_col = area.Column
            n_cols = area.Columns.Count
            total_width = sum(
                ws.Columns(first_col + i).ColumnWidth for i in range(n_cols)
            )

            scratch_col.ColumnWidth = total_width
            scratch_cell = ws.Cells(measurement_row, scratch_col_index)
            scratch_cell.WrapText = True
            scratch_cell.Value = anchor.Value
            # Copy font properties so the measurement reflects the real cell.
            scratch_cell.Font.Name   = anchor.Font.Name
            scratch_cell.Font.Size   = anchor.Font.Size
            scratch_cell.Font.Bold   = anchor.Font.Bold
            scratch_cell.Font.Italic = anchor.Font.Italic
            scratch_cell.EntireRow.AutoFit()
            needed = scratch_cell.RowHeight

            anchor_row = ws.Rows(area.Row)
            if needed > anchor_row.RowHeight:
                anchor_row.RowHeight = needed
                adjusted += 1
    finally:
        scratch_col.Clear()
        scratch_col.ColumnWidth = original_width

    return adjusted

_FILENAME_BAD = re.compile(r'[/\\:*?"<>|]+')

def _safe_filename(s: str) -> str:
    cleaned = _FILENAME_BAD.sub("_", s).strip()
    return cleaned or "sheet"

def xlsx_to_pdf(xlsx_path: Path, out_dir: Path) -> dict[str, Path]:
    """Convert an .xlsx to one .pdf **per worksheet** using Excel COM.

    Returns ``{sheet_name: pdf_path}`` so downstream stages can attribute
    extracted questions to the right sheet (workbooks routinely contain
    several questionnaires on separate sheets). Each sheet's print area is
    restricted to ``{PRINT_COL_START}:{PRINT_COL_END}`` (default B:AN),
    scaled to one page wide, and has merged-wrap-text row heights expanded.
    Windows + Excel only.
    """
    try:
        import pythoncom
        import win32com.client
    except ImportError as e:
        raise RuntimeError(
            "pywin32 (win32com) is required for the xlsx→pdf step. "
            "Install with `pip install pywin32` on Windows with Excel."
        ) from e

    out_dir.mkdir(parents=True, exist_ok=True)
    src_path = xlsx_path.resolve()

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        wb = excel.Workbooks.Open(str(src_path), ReadOnly=True, UpdateLinks=0)

        # Snapshot last row per sheet BEFORE any mutation so the scratch-column
        # writes inside _autofit_merged_rows can't inflate it.
        sheet_last_rows: dict[str, int] = {
            ws.Name: _sheet_last_row(ws) for ws in wb.Worksheets
        }

        out: dict[str, Path] = {}
        used_stems: set[str] = set()
        for ws in wb.Worksheets:
            sheet_name = ws.Name
            last_row = sheet_last_rows[sheet_name]
            n_adjusted = _autofit_merged_rows(ws, last_row=last_row)
            if n_adjusted:
                print(f"    auto-expanded {n_adjusted} merged row(s) on '{sheet_name}'")

            ws.PageSetup.PrintArea = (
                f"${PRINT_COL_START}$1:${PRINT_COL_END}${last_row}"
            )
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False

            base = f"{xlsx_path.stem}__{_safe_filename(sheet_name)}"
            stem = base
            i = 2
            while stem.lower() in used_stems:
                stem = f"{base}_{i}"
                i += 1
            used_stems.add(stem.lower())

            sheet_pdf = (out_dir / f"{stem}.pdf").resolve()
            ws.ExportAsFixedFormat(XL_TYPE_PDF, str(sheet_pdf))
            if not sheet_pdf.exists():
                raise RuntimeError(
                    f"Excel did not produce {sheet_pdf} for sheet '{sheet_name}'."
                )
            out[sheet_name] = sheet_pdf

        return out
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        finally:
            if excel is not None:
                excel.Quit()
            pythoncom.CoUninitialize()

# ─── pdf → PNG pages ────────────────────────────────────────────────────────
def render_pdf_pages(pdf_path: Path, dpi: int = 200) -> list[Image.Image]:
    """Render every page to a PIL Image. Works on read-only / locked PDFs."""
    scale = dpi / 72.0
    pdf = pdfium.PdfDocument(str(pdf_path))
    try:
        return [
            pdf[i].render(scale=scale).to_pil().convert("RGB")
            for i in range(len(pdf))
        ]
    finally:
        pdf.close()

def image_to_data_url(img: Image.Image, max_side: int = 2000) -> str:
    """PNG-encode and base64-wrap a PIL image. Downscale if a side exceeds max_side."""
    w, h = img.size
    if max(w, h) > max_side:
        ratio = max_side / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()

# ─── Azure OpenAI extraction ────────────────────────────────────────────────
SYSTEM_PROMPT = (
    "You extract questionnaire content from images of PDF pages. "
    "Each image is one page of a multi-page questionnaire. The user may send "
    "two consecutive pages at once so that questions spanning the page break "
    "are fully visible. "
    "Identify every question that has a printed question id (e.g. 'Q1', "
    "'Q12.a', '3.2'), and return its stem text plus a list of sub_questions, "
    "each carrying its own selection and answer.\n"
    "Filling sub_questions:\n"
    "- A question often bundles several answerable parts. Return ONE "
    "  sub_question per part, in the order they appear — never collapse them "
    "  into one field and never pick just one part.\n"
    "- Multi-part questions ('a) …', 'b) …'): one sub_question per part, with "
    "  that part's text in `prompt` and its own `answer`.\n"
    "- Branch / option rows (e.g. 'Settling distributor' vs 'Non-settling "
    "  distributor'): one sub_question per option. Put the option text in "
    "  `option_label`, any follow-up prompt (e.g. 'Who is the custodian?') in "
    "  `prompt`, the chosen YES/NO (or ticked) value in `selection`, and the "
    "  free text the respondent wrote in `answer`.\n"
    "- A plain question with a single answer box and no distinct sub-parts: "
    "  return exactly one sub_question with empty `option_label`, `prompt`, "
    "  and `selection`, and the respondent's text in `answer`.\n"
    "- Labelled-section questions — a short title stem followed by one or more "
    "  labelled sections down the page (e.g. stem 'Relevant Authorised "
    "  Activities' with sections 'For HK, please specify license type' and "
    "  'For Singapore, please specify license type'): keep ONLY the short title "
    "  in `question`, and emit ONE sub_question per labelled section in document "
    "  order. Put the section's label in `prompt` and the respondent's entry "
    "  for that section in `answer`. Emit a sub_question for every section even "
    "  if its answer is blank. Large vertical whitespace between sections does "
    "  NOT end the question — it ends only when the next printed question id "
    "  appears.\n"
    "Field rules:\n"
    "- `selection` is only for a chosen control (dropdown like YES/NO, a "
    "  ticked checkbox/radio, a highlighted option). Never put free text in "
    "  `selection`.\n"
    "- `answer` is only what the respondent wrote/typed. If an answer box is "
    "  empty, use an empty string — do not guess or copy the prompt.\n"
    "- Keep `question` to just the stem printed next to the id; do not fold "
    "  sub-prompts or option labels into it.\n"
    "Telling answers from template/placeholder text:\n"
    "- Styling (italic, grey) does NOT tell you whether text is an instruction "
    "  or a real answer — a respondent's typed answer can also be italic or "
    "  greyed. Decide by meaning, not styling.\n"
    "- Template / instruction / placeholder text is generic guidance addressed "
    "  to the respondent (e.g. 'For X, please specify ...', 'Please detail ...', "
    "  'if available', 'e.g. ...', 'etc.', '(if applicable)'). It is NEVER an "
    "  answer: put a section's instruction text in `prompt`, and if a section "
    "  has only such text and no concrete entry, leave `answer` empty.\n"
    "- A real answer is specific to this respondent: concrete names, codes, "
    "  numbers, or entity-specific lists (e.g. 'Authorized Institution - "
    "  HKMA'). Capture those in `answer`.\n"
    "Coverage rules:\n"
    "- Only return a question if its id is fully visible. If an id is on the "
    "  first image and its body/answers continue on the second image, return "
    "  the complete merged item. If an id starts at the very bottom and its "
    "  body is cut off (not visible on either image in this window), skip it — "
    "  another window will see it in full.\n"
    "- Preserve newlines inside long prompts or answers. When one section's "
    "  answer is a multi-line list (one entry per line), keep it as a single "
    "  `answer` with '\\n' separators — do not split it into multiple "
    "  sub_questions.\n"
    "- Do not invent ids, questions, selections, or answers. Do not summarise.\n"
    "Worked example (labelled-section question):\n"
    "Page shows id 'Q3', stem title 'Relevant Authorised Activities'. A section "
    "'For HK, please specify license type' is filled with four lines: "
    "'Authorized Institution - HKMA', 'Dealing in Securities - SFC', 'Advising "
    "on Securities - SFC', 'Advising on Securities - SFC'. Further down, after "
    "a large gap, a section 'For Singapore, please specify license type' shows "
    "only the placeholder 'Please detail the license type and conditions.' with "
    "nothing filled in. Correct output:\n"
    '{"questions":[{"question_id":"Q3","question":"Relevant Authorised '
    'Activities","sub_questions":[{"option_label":"","prompt":"For HK, please '
    'specify license type","selection":"","answer":"Authorized Institution - '
    'HKMA\\nDealing in Securities - SFC\\nAdvising on Securities - SFC\\nAdvising '
    'on Securities - SFC"},{"option_label":"","prompt":"For Singapore, please '
    'specify license type","selection":"","answer":""}]}]}\n'
    "The title stays in `question`, each section is one sub_question, the HK "
    "answer keeps its newlines, and the Singapore placeholder is left out of "
    "`answer`.\n"
)

def build_llm(model_kwargs: dict | None = None) -> AzureChatOpenAI:
    """Construct the underlying Azure ChatOpenAI client. Call this **once**
    per process — it opens an HTTP client and configures the deployment."""
    endpoint   = os.environ.get("AZURE_OPENAI_ENDPOINT")
    api_key    = os.environ.get("AZURE_OPENAI_API_KEY")
    deployment = os.environ.get("AZURE_OPENAI_DEPLOYMENT")
    api_ver    = os.environ.get("AZURE_OPENAI_API_VERSION", DEFAULT_API_VERSION)
    missing = [n for n, v in [
        ("AZURE_OPENAI_ENDPOINT", endpoint),
        ("AZURE_OPENAI_API_KEY", api_key),
        ("AZURE_OPENAI_DEPLOYMENT", deployment),
    ] if not v]
    if missing:
        raise RuntimeError(
            "Missing required environment variables: " + ", ".join(missing)
        )
    return AzureChatOpenAI(
        azure_endpoint=endpoint,
        api_key=api_key,
        azure_deployment=deployment,
        api_version=api_ver,
        temperature=0,
        # We do our own retry (call_with_retry) so the langchain layer
        # shouldn't add hidden retries on top.
        max_retries=0,
        **(model_kwargs or {}),
    )

def build_structured_llm(
    model_kwargs: dict | None = None,
) -> Runnable:
    """Construct the structured-output runnable used by every window.

    Wraps ``build_llm()`` with ``with_structured_output(QAList)`` exactly
    once so neither the underlying HTTP client nor the structured-output
    binding gets re-created per LLM call.
    """
    return build_llm(model_kwargs).with_structured_output(QAList)

T = TypeVar("T")

def call_with_retry(
    fn: Callable[[], T],
    *,
    max_attempts: int = DEFAULT_MAX_ATTEMPTS,
    base_delay: float = DEFAULT_RETRY_BASE_DELAY,
    label: str = "call",
) -> T:
    """Run ``fn()`` with exponential-backoff retry on **any** exception.

    Retries on Pydantic ValidationError (LLM returned malformed structured
    output), 404 / 429 / 5xx HTTP errors from the API, transient network
    failures, and anything else. The final attempt's exception propagates.
    Delays follow 2s, 4s, 8s, ... by default.
    """
    for attempt in range(1, max_attempts + 1):
        try:
            return fn()
        except Exception as e:
            if attempt >= max_attempts:
                raise
            delay = base_delay * (2 ** (attempt - 1))
            print(
                f"      {label} attempt {attempt}/{max_attempts} failed "
                f"({type(e).__name__}: {e}); retrying in {delay:.1f}s",
                file=sys.stderr,
            )
            time.sleep(delay)
    # Unreachable: the loop either returns or raises.
    raise RuntimeError("call_with_retry exhausted attempts without raising")

def extract_from_window(
    structured_llm: Runnable,
    page_images: list[Image.Image],
    page_numbers: list[int],
) -> list[dict]:
    """Run the vision model on a window of consecutive pages and return parsed
    items. ``structured_llm`` is the pre-built ``with_structured_output``
    runnable from ``build_structured_llm()`` — it is reused across every
    window so we don't re-initialise the client. Each returned dict carries
    ``source_pages`` so we know which window produced it."""
    user_content: list[dict] = [{
        "type": "text",
        "text": (
            f"This window contains page(s) {', '.join(map(str, page_numbers))} "
            f"of the questionnaire (in order). Extract every question per the "
            f"rules in the system message."
        ),
    }]
    for img in page_images:
        user_content.append({
            "type": "image_url",
            "image_url": {"url": image_to_data_url(img), "detail": "high"},
        })

    result: QAList = structured_llm.invoke([
        SystemMessage(content=SYSTEM_PROMPT),
        HumanMessage(content=user_content),
    ])
    items: list[dict] = []
    for q in result.questions:
        qid = q.question_id.strip()
        if not qid:
            continue
        sub_questions = [
            {
                "option_label":  sq.option_label.strip(),
                "prompt":        sq.prompt.strip(),
                "selection":     sq.selection.strip(),
                "answer":        sq.answer.strip(),
                "answer_source": "llm",
            }
            for sq in q.sub_questions
        ]
        items.append({
            "question_id":   qid,
            "question":      q.question.strip(),
            "sub_questions": sub_questions,
            "source_pages":  list(page_numbers),
        })
    return items

# ─── Sliding window + dedup ─────────────────────────────────────────────────
def iter_windows(n_pages: int, window: int, stride: int) -> Iterable[list[int]]:
    """1-indexed page-number windows. The final window is always anchored at the
    last page so it never gets dropped."""
    if n_pages == 0:
        return
    if n_pages <= window:
        yield list(range(1, n_pages + 1))
        return
    starts = list(range(1, n_pages - window + 2, stride))
    if starts[-1] != n_pages - window + 1:
        starts.append(n_pages - window + 1)
    for s in starts:
        yield list(range(s, s + window))

def _completeness(item: dict) -> tuple[int, int, int]:
    """Score used to pick the better duplicate. Prefer the variant with more
    filled sub-questions (a non-empty selection or answer), then more
    sub-questions, then more total text."""
    subs = item.get("sub_questions", [])
    n_filled = sum(1 for s in subs if s.get("selection") or s.get("answer"))
    text_len = len(item.get("question", "")) + sum(
        len(s.get("option_label", "")) + len(s.get("prompt", ""))
        + len(s.get("selection", "")) + len(s.get("answer", ""))
        for s in subs
    )
    return (n_filled, len(subs), text_len)

def dedupe_by_id(items: list[dict]) -> list[dict]:
    """Dedup keyed by (sheet, question_id): the same Q-id can appear on two
    different sheets of the same workbook and must stay separate."""
    best: dict[tuple[str, str], dict] = {}
    for it in items:
        key = (it.get("sheet", ""), it["question_id"])
        if key not in best:
            best[key] = it
            continue
        if _completeness(it) > _completeness(best[key]):
            merged_pages = sorted(
                set(best[key]["source_pages"]) | set(it["source_pages"])
            )
            it = {**it, "source_pages": merged_pages}
            best[key] = it
        else:
            best[key]["source_pages"] = sorted(
                set(best[key]["source_pages"]) | set(it["source_pages"])
            )

    def sort_key(k: tuple[str, str]):
        sheet, qid = k
        nums = [int(n) for n in re.findall(r"\d+", qid)]
        return (sheet, nums or [10**9], qid)

    return [best[k] for k in sorted(best, key=sort_key)]

# ─── xlsx reconcile (truncation safety net) ────────────────────────────────
def _normalise(s: str) -> str:
    return " ".join(s.split()).casefold()

def _collect_xlsx_answers(xlsx_path: Path) -> dict[tuple[str, str], list[str]]:
    """(sheet_name, qid) → ordered list of non-empty answer-cell texts, by
    scanning the source xlsx with openpyxl. Same notion of "answer cell" as
    parse.py: grey/yellow fill in the B:AO content band, attributed to the
    most recent Q-id on that sheet. Returned in document order so they can be
    matched positionally to a question's sub_questions."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    answers: dict[tuple[str, str], list[str]] = {}

    for ws in wb.worksheets:
        sheet_name = ws.title
        merged = build_merged_lookup(ws)
        current_qid: str | None = None
        seen_anchors: set = set()

        for row in range(1, ws.max_row + 1):
            for col in range(CONTENT_START_COL, CONTENT_END_COL + 1):
                anchor = merged.get((row, col), (row, col))
                if anchor in seen_anchors:
                    continue
                seen_anchors.add(anchor)

                cell = ws.cell(anchor[0], anchor[1])
                val = effective_value(ws, row, col, merged)
                text = "" if val is None else str(val).strip()

                m = Q_PATTERN.match(text) if text else None
                if m:
                    current_qid = m.group(0)
                    continue
                if current_qid is None:
                    continue
                if not text:
                    continue
                if classify_fill(cell) in ANSWER_FILLS:
                    key = (sheet_name, current_qid)
                    answers.setdefault(key, []).append(text)
    return answers

def _looks_truncated(llm_ans: str, truth: str) -> bool:
    """True when the LLM answer is a prefix of the xlsx text and the xlsx has
    materially more content — the signature of a vision-model truncation."""
    truth_n = _normalise(truth)
    llm_n   = _normalise(llm_ans)
    return truth_n.startswith(llm_n) and len(truth_n) > len(llm_n) + 5

def reconcile_with_xlsx(
    items: list[dict], xlsx_path: Path
) -> tuple[list[dict], int]:
    """Repair LLM sub-question answers that look truncated (or were missed)
    using the full text from the source xlsx. Matches by (sheet, question_id),
    then positionally between the xlsx's ordered answer cells and the question's
    sub_questions. Conservative: only acts when the counts line up, so an
    ambiguous mismatch leaves the LLM output untouched. Modifies items in place
    and returns them, plus the count of substitutions made."""
    answers = _collect_xlsx_answers(xlsx_path)
    n_reconciled = 0

    for item in items:
        truths = answers.get((item.get("sheet", ""), item["question_id"]))
        if not truths:
            continue
        subs = item.setdefault("sub_questions", [])
        answered = [s for s in subs if s.get("answer")]

        if answered and len(answered) == len(truths):
            # Fix truncations on the answers the LLM did capture.
            for s, truth in zip(answered, truths):
                if _looks_truncated(s["answer"], truth):
                    s["answer"] = truth
                    s["answer_source"] = "xlsx_reconciled"
                    n_reconciled += 1
        elif not answered and subs and len(subs) == len(truths):
            # LLM saw the sub-questions but missed every answer; fill in order.
            for s, truth in zip(subs, truths):
                s["answer"] = truth
                s["answer_source"] = "xlsx_reconciled"
                n_reconciled += 1
        elif not answered and len(truths) == 1 and len(subs) <= 1:
            # Simple single-answer question the LLM left blank (or split into
            # no sub-questions at all).
            if not subs:
                subs.append({
                    "option_label": "", "prompt": "", "selection": "",
                    "answer": "", "answer_source": "llm",
                })
            subs[0]["answer"] = truths[0]
            subs[0]["answer_source"] = "xlsx_reconciled"
            n_reconciled += 1
        # else: ambiguous count mismatch — leave the LLM output as-is.

    return items, n_reconciled

# ─── Orchestration ──────────────────────────────────────────────────────────
@dataclass
class ParseResult:
    file_name: str
    items: list[dict]

def parse_one(
    src: Path,
    out_dir: Path,
    structured_llm: Runnable,
    dpi: int,
    window: int,
    stride: int,
    keep_pngs: bool,
    max_attempts: int,
) -> ParseResult:
    """Process a single xlsx (or pdf) end-to-end.

    For xlsx: every worksheet is exported to its own PDF, processed
    separately, and items carry the originating sheet name. For pdf: the
    file is processed as a single 'sheet' named after the file stem.
    """
    # sheet_name → pdf_path
    sheet_pdfs: dict[str, Path]
    xlsx_source: Path | None = None
    if src.suffix.lower() == ".pdf":
        sheet_pdfs = {src.stem: src}
    elif src.suffix.lower() in {".xlsx", ".xlsm"}:
        print(f"  → converting {src.name} to PDF via Excel COM (per sheet)")
        sheet_pdfs = xlsx_to_pdf(src, out_dir)
        xlsx_source = src
        print(f"    {len(sheet_pdfs)} sheet(s): {', '.join(sheet_pdfs)}")
    else:
        raise ValueError(f"Unsupported input type: {src.suffix}")

    all_items: list[dict] = []
    for sheet_name, pdf_path in sheet_pdfs.items():
        print(f"  → sheet '{sheet_name}': rendering {pdf_path.name} at {dpi} dpi")
        pages = render_pdf_pages(pdf_path, dpi=dpi)
        n = len(pages)
        print(f"    {n} page(s)")

        if keep_pngs:
            png_dir = out_dir / f"{pdf_path.stem}_pages"
            png_dir.mkdir(parents=True, exist_ok=True)
            for i, img in enumerate(pages, 1):
                img.save(png_dir / f"page_{i:03d}.png")

        for win in iter_windows(n, window=window, stride=stride):
            print(f"    LLM call on pages {win}")
            win_imgs = [pages[p - 1] for p in win]
            try:
                items = call_with_retry(
                    lambda: extract_from_window(structured_llm, win_imgs, win),
                    max_attempts=max_attempts,
                    label=f"sheet '{sheet_name}' pages {win}",
                )
            except Exception as e:
                print(
                    f"      ! gave up on sheet '{sheet_name}' window {win} "
                    f"after {max_attempts} attempts: {e}",
                    file=sys.stderr,
                )
                continue
            for it in items:
                it["sheet"] = sheet_name
            all_items.extend(items)

    merged = dedupe_by_id(all_items)
    print(f"  extracted {len(merged)} unique question(s) across all sheets")

    if xlsx_source is not None:
        try:
            merged, n_rec = reconcile_with_xlsx(merged, xlsx_source)
            if n_rec:
                print(f"  reconciled {n_rec} answer(s) from xlsx source text")
        except Exception as e:
            print(f"  ! xlsx reconcile failed: {e}", file=sys.stderr)

    return ParseResult(file_name=src.stem, items=merged)

# ─── Writers ────────────────────────────────────────────────────────────────
# One row per sub-question. Column names match parse.py's flattened schema
# (question / option_label / prompt / side_answer / answer) so the downstream
# analysers (analyze.py, analyze_llm_xlsx.py) read this output unchanged.
XLSX_FIELDS = [
    "file_name", "sheet", "question_id", "sub_idx", "question",
    "option_label", "prompt", "side_answer", "answer",
    "answer_source", "source_pages",
]

def write_json(result: ParseResult, path: Path) -> None:
    path.write_text(
        json.dumps(
            {"file_name": result.file_name, "questions": result.items},
            indent=2, ensure_ascii=False,
        ),
        encoding="utf-8",
    )

def _row_question_text(stem: str, sub: dict) -> str:
    """Per-row question text: the stem, with this sub-question's prompt appended
    so multi-part questions stay distinguishable when flattened."""
    prompt = sub.get("prompt", "")
    if prompt and prompt != stem:
        return f"{stem} — {prompt}" if stem else prompt
    return stem

def write_xlsx(results: list[ParseResult], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "questions"
    ws.append(XLSX_FIELDS)
    for r in results:
        for it in r.items:
            stem = it.get("question", "")
            pages = ", ".join(map(str, it.get("source_pages", [])))
            # A question with no sub_questions still emits one (empty) row.
            subs = it.get("sub_questions") or [{}]
            for idx, sub in enumerate(subs):
                ws.append([
                    r.file_name,
                    it.get("sheet", ""),
                    it["question_id"],
                    idx,
                    _row_question_text(stem, sub),
                    sub.get("option_label", ""),
                    sub.get("prompt", ""),
                    sub.get("selection", ""),
                    sub.get("answer", ""),
                    sub.get("answer_source", "llm"),
                    pages,
                ])
    wb.save(path)

# ─── CLI ────────────────────────────────────────────────────────────────────
def main() -> None:
    ap = argparse.ArgumentParser(
        description="PDF-vision questionnaire parser (Azure OpenAI via langchain).",
    )
    ap.add_argument("target", help="Path to .xlsx/.pdf file OR directory of them")
    ap.add_argument("--output-dir", default="./output",
                    help="Output directory (default: ./output)")
    ap.add_argument("--dpi", type=int, default=200,
                    help="Render DPI for PDF→PNG (default 200)")
    ap.add_argument("--window", type=int, default=2,
                    help="Pages per LLM call (default 2 = sliding pair)")
    ap.add_argument("--stride", type=int, default=1,
                    help="Step between window starts (default 1 = overlap)")
    ap.add_argument("--keep-pngs", action="store_true",
                    help="Also save the rendered page PNGs alongside the PDF")
    ap.add_argument("--max-attempts", type=int, default=DEFAULT_MAX_ATTEMPTS,
                    help=f"Max LLM attempts per window before giving up "
                         f"(default {DEFAULT_MAX_ATTEMPTS}). Custom retry — "
                         f"langchain's own retry is disabled.")
    args = ap.parse_args()

    target  = Path(args.target)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if target.is_dir():
        files = sorted(
            p for p in target.iterdir()
            if p.suffix.lower() in {".xlsx", ".xlsm", ".pdf"}
            and not p.name.startswith("~$")
        )
    else:
        files = [target]
    if not files:
        print(f"No xlsx/pdf inputs found in {target}", file=sys.stderr)
        sys.exit(1)

    # Build the model and the structured-output wrapper exactly once for the
    # whole run; reuse across every file, sheet, and window.
    structured_llm = build_structured_llm()

    results: list[ParseResult] = []
    for fp in files:
        print(f"→ {fp.name}")
        try:
            res = parse_one(
                fp, out_dir, structured_llm,
                dpi=args.dpi, window=args.window, stride=args.stride,
                keep_pngs=args.keep_pngs, max_attempts=args.max_attempts,
            )
        except Exception as e:
            print(f"  ✗ failed: {e}", file=sys.stderr)
            continue
        write_json(res, out_dir / f"{fp.stem}.json")
        results.append(res)

    if not results:
        print("No outputs produced.", file=sys.stderr)
        sys.exit(2)

    if target.is_dir():
        from datetime import date
        xlsx_path = out_dir / f"{target.name}_pdf_{date.today().isoformat()}.xlsx"
    else:
        xlsx_path = out_dir / f"{results[0].file_name}.xlsx"
    write_xlsx(results, xlsx_path)
    print(f"\n✓ Wrote {xlsx_path}")

if __name__ == "__main__":
    main()
