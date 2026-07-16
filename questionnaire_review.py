"""Cross-questionnaire audit of long-format questionnaire xlsx files.

Input: one or more .xlsx files, each with a sheet ``Questionnaire_Long`` and
columns::

    Source_file | Section | question_id | question | answer

(the layout produced by generate_questionnaires.py). Each file is one
questionnaire; question_ids are NOT trusted for matching (the same question can
be KYC_010 in one file and KYC_012 in another) — questions are matched across
files by text similarity, the same as kyd_review.py.

Two things are produced in one xlsx (+ a matching self-contained HTML heat map):

1. **Summary** (first tab) — one row per questionnaire: number of sections,
   number of questions, count/percentage of missing (blank) answers, and
   count/percentage of declared not-applicable answers (na / N/A / not
   applicable / nil / none).

2. **Cross-questionnaire checks** — question wording drift, questions
   present in only a few / missing from a file, and answer outliers.

Answers are free text. They are handled in two disjoint ways:

  * **Pure numbers** (``45``, ``1,250,000``, ``~50``, ``30%``) are compared
    numerically (robust median / MAD). They are NEVER sent to the embedding
    model — an embedding of "15000000" vs "42000000" encodes the token shape,
    not the magnitude, so it would cluster unrelated figures and hide the
    outlier that matters.
  * **Everything else** (real free text) is embedded and grouped by cosine
    similarity, so paraphrases group together and a lone divergent answer
    surfaces as an outlier. Falls back to difflib when embeddings are off.

    python questionnaire_review.py "kyc demo"/*.xlsx -o output/q_review.html
    python questionnaire_review.py --dir //server/share/batch1 --no-embed

Handles long / UNC Windows paths (extended-length ``\\\\?\\`` form) and merges
multiple input files/folders into a single result.
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

# Reuse the proven machinery from kyd_review — question matching, wording-variant
# classification, embedding cache, long-path shim, numeric parse, and both
# renderers. Only the reader, the free-text answer engine, the Summary sheet and
# the (supplementary) LLM anomaly layer are new here. kyd_review is left intact.
import kyd_review_compat as kyd
from kyd_review_compat import (
    OK, NOTE, WARN, MISSING, OUTLIER,
    SPECIFIC_SHARE, REQUIRED_SHARE, MINORITY_SHARE,
    norm, ratio, canon, answer_class, _num, q_signature,
    cluster_questions, cluster_title, check_variants, _flag,
    embed_signatures, load_config, _winlong, render_html, render_xlsx,
)

SHEET = "Questionnaire_Long"
NEEDED = ("source_file", "section", "question_id", "question", "answer")


# --- reading long-format xlsx ------------------------------------------------

def _open_wb(path: Path):
    from openpyxl import load_workbook
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except (FileNotFoundError, OSError):
        return load_workbook(str(_winlong(path)), read_only=True, data_only=True)


def _cell(v) -> str:
    return "" if v is None else str(v).strip()


def read_questionnaire(path: Path) -> dict:
    """Parse one xlsx into kyd_review's file shape.

    Each row becomes a question carrying its free-text answer directly (no
    sub_questions — this format has exactly one answer per question).
    """
    wb = _open_wb(path)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = [_cell(h).lower() for h in next(rows, [])]
    idx = {name: header.index(name) for name in NEEDED if name in header}
    missing_cols = [c for c in NEEDED if c not in idx]
    if missing_cols:
        raise ValueError(f"{path.name}: sheet '{SHEET}' missing columns {missing_cols}")

    questions, source = [], ""
    for r in rows:
        get = lambda name: _cell(r[idx[name]]) if idx[name] < len(r) else ""  # noqa: E731
        section, qid, question, answer = get("section"), get("question_id"), get("question"), get("answer")
        source = source or get("source_file")
        if not (question or answer or qid):
            continue  # skip fully blank rows
        questions.append({"question_id": qid, "question": question, "section": section,
                          "answer": answer, "sub_questions": []})
    wb.close()
    return {"raw_name": source or path.stem, "questions": questions}


def load_files(entries: list[tuple[Path, str | None]]) -> list[dict]:
    """entries: (path, folder_stem_or_None); folder set only when >1 --dir given."""
    raw = []
    for p, folder in sorted(entries, key=lambda e: (e[1] or "", e[0].name)):
        d = read_questionnaire(p)
        d["folder"] = folder or ""
        raw.append(d)
    dupes = Counter(f["raw_name"] for f in raw)
    for f in raw:
        f["name"] = f["raw_name"] if dupes[f["raw_name"]] == 1 else f"{f['raw_name']} ({f['folder']})"
    return raw


def file_stats(f: dict) -> dict:
    qs = f["questions"]
    n = len(qs)
    sections = {q["section"] for q in qs if q["section"]}
    cls = Counter(answer_class(q["answer"]) for q in qs)
    blank, na = cls.get("blank", 0), cls.get("na", 0)
    pct = lambda c: round(100 * c / n, 1) if n else 0.0  # noqa: E731
    return {"name": f["name"], "sections": len(sections), "questions": n,
            "answered": cls.get("text", 0), "missing": blank, "missing_pct": pct(blank),
            "na": na, "na_pct": pct(na)}


# --- free-text / numeric answer engine ---------------------------------------

def _unit(vec):
    import numpy as np
    v = np.asarray(vec, dtype=float)
    nrm = np.linalg.norm(v)
    return v / nrm if nrm else v


def _answer_sim(a: str, b: str, avecs: dict | None) -> float:
    """Cosine of two answers when both embedded; else difflib on canon text.

    Pure-number answers are absent from ``avecs`` by construction, so any pair
    involving one degrades to the lexical ratio — they are meant to be judged by
    the numeric path, not here.
    """
    if avecs is not None:
        ka, kb = norm(a), norm(b)
        if ka in avecs and kb in avecs:
            return float(_unit(avecs[ka]) @ _unit(avecs[kb]))
    return ratio(canon(a), canon(b))


def check_answers(entries: list[dict], findings: list, ci: int,
                  avecs: dict | None, cos_thr: float) -> None:
    """Blank / declared-N/A / numeric-outlier / free-text-outlier on one question."""
    n = len(entries)
    if n < 3:
        return
    cls = lambda e: answer_class(e["answer"])  # noqa: E731
    answered = [e for e in entries if cls(e) != "blank"]
    filled = [e for e in entries if cls(e) == "text"]  # substantive (incl. numbers)

    if answered and len(answered) / n >= REQUIRED_SHARE:
        for e in entries:
            if cls(e) == "blank":
                _flag(e, findings, ci, MISSING, "missing answer",
                      f"Answer left blank while {len(answered)}/{n} firms answered.")
    if answered and len(filled) / len(answered) >= REQUIRED_SHARE:
        for e in answered:
            if cls(e) == "na":
                _flag(e, findings, ci, MISSING, "declared N/A",
                      f"Firm answered “{e['answer']}” while {len(filled)}/{len(answered)} peers gave a "
                      f"substantive answer — the not-applicable claim should be challenged.")
    if len(filled) < 5:
        return

    # numeric answers: compare magnitudes, never embeddings
    nums = [(_num(e["answer"]), e) for e in filled]
    import statistics
    if sum(1 for v, _ in nums if v is not None) / len(filled) >= 0.8:
        vals = [v for v, _ in nums if v is not None]
        med = statistics.median(vals)
        mad = statistics.median([abs(v - med) for v in vals])
        for v, e in nums:
            if v is None:
                continue
            z = 0.6745 * abs(v - med) / mad if mad else (abs(v - med) / (abs(med) or 1))
            if z > 3.5:
                _flag(e, findings, ci, OUTLIER, "numeric outlier",
                      f"Value {e['answer']} is far from the peer median of {med:g}.")
        return

    # free text: greedy semantic groups (cosine if embedded, else difflib),
    # a small minority group diverging from the common wording => outlier
    groups: list[list[dict]] = []
    for e in filled:
        g = next((g for g in groups if _answer_sim(e["answer"], g[0]["answer"], avecs) >= cos_thr), None)
        (g.append(e) if g is not None else groups.append([e]))
    top = max(len(g) for g in groups)
    if top / len(filled) >= 0.6:
        for g in groups:
            if len(g) / len(filled) <= MINORITY_SHARE and len(g) < top:
                for e in g:
                    _flag(e, findings, ci, OUTLIER, "free-text outlier",
                          f"Answer diverges from the wording used by {top}/{len(filled)} firms: "
                          f"“{e['answer'][:100]}”")


# --- analysis (builds the same report dict kyd_review's renderers consume) ----

def analyze(files: list[dict], multi_folder: bool, qvecs, avecs,
            threshold, cos_thr: float) -> dict:
    n_files = len(files)
    clusters = cluster_questions(files, qvecs, threshold)
    findings: list[dict] = []
    out_clusters = []
    for ci, c in enumerate(clusters):
        coverage = len(c["members"])
        specific = coverage / n_files <= SPECIFIC_SHARE
        if specific:
            for fn in c["members"]:
                findings.append(dict(file=fn, cluster=ci, level=WARN, kind="questionnaire-specific",
                                     message=f"Question appears in only {coverage}/{n_files} questionnaires."))
        else:
            for f in files:
                if f["name"] not in c["members"]:
                    findings.append(dict(file=f["name"], cluster=ci, level=MISSING, kind="question missing",
                                         message=f"Question present in {coverage}/{n_files} questionnaires but absent here."))
        variants = check_variants(c, findings, ci)
        entries = [{"file": fn, "selection": "", "answer": q.get("answer", ""), "flags": []}
                   for fn, q in c["members"].items()]
        check_answers(entries, findings, ci, avecs, cos_thr)
        out_clusters.append({
            "title": cluster_title(c),
            "ids": {fn: q["question_id"] for fn, q in c["members"].items()},
            "coverage": coverage, "specific": specific, "variants": variants,
            "slots": [{"label": "answer", "entries": entries}],
        })

    # cells / qvar / sort — identical shape to kyd_review.analyze so the renderers work
    fnames = [f["name"] for f in files]
    per_cell: dict = {}
    for g in findings:
        per_cell.setdefault((g["file"], g["cluster"]), []).append(g)
    cells = {fn: [] for fn in fnames}
    for fn in fnames:
        for ci, c in enumerate(out_clusters):
            if fn not in c["ids"] and c["specific"]:
                cells[fn].append(-1)
            else:
                cells[fn].append(max((g["level"] for g in per_cell.get((fn, ci), [])), default=OK))
    findings.sort(key=lambda g: (-g["level"], g["cluster"], g["file"]))

    kindcode = {"majority": 0, "formatting": 1, "substantive": 2}
    qvar = []
    for c in out_clusters:
        fmap = {}
        for v in c["variants"]:
            for fn in v["files"]:
                fmap[fn] = kindcode.get(v.get("kind"), 0)
        qvar.append([fmap.get(fn, -1) for fn in fnames])

    folders = {f["name"]: f.get("folder", "") for f in files}
    return {"files": fnames, "clusters": out_clusters, "cells": cells, "findings": findings,
            "folders": folders, "multi_folder": multi_folder, "qvar": qvar}


# --- LLM anomaly layer (Layer 1: generic, peer-context, embedding-compressed) -
#
# Supplement to the rule-based review above, NOT a replacement. No domain rubric:
# the model is shown one canonical question plus its peer answers and asked which
# answers are anomalous / non-answers / self-inconsistent / implausible. The norm
# is learned from the corpus, so this generalises to any questionnaire domain.
#
# Scale: embeddings pre-group the answers; only group representatives (+ counts)
# and the minority answers are sent, so the prompt stays bounded no matter how
# many firms answered. Pure numbers are never sent — the numeric path owns them.

LLM_PROMPT_VERSION = "l1-2026-07-07"
MIN_PEERS = 3          # need at least this many text answers to define a norm
LLM_COLORS = ["#5cb85c", "#9ec5f4", "#fab219", "#d03b3b"]
LLM_NAMES = {0: "no anomaly", 1: "minor", 2: "anomaly / non-answer",
             3: "implausible / inconsistent"}
KIND_SEV = {"outlier": 2, "non_answer": 2, "inconsistent": 3, "implausible": 3}

LLM_SYSTEM = (
    "You audit questionnaire answers. For ONE question you are given the peer "
    "answers from several respondents (frequent answers are deduplicated with a "
    "count). Using only general reasoning about the question and the peer set — "
    "no outside domain assumptions — identify answers that are anomalous relative "
    "to peers, non-answers or evasive, internally inconsistent, or implausible "
    "for what the question asks. Reply ONLY with a JSON object of the form "
    '{"flagged":[{"id":<int>,"kind":"outlier|non_answer|inconsistent|implausible",'
    '"rationale":"<short reason>"}]}. Use the ids shown. Empty list if all fine.'
)


def _compress_slot(title: str, entries: list[dict], avecs, cos_thr: float,
                   max_answers: int) -> dict | None:
    """Bounded peer-context payload for one question.

    Groups text answers by similarity, sends group representatives (+ counts)
    for the majority and every minority answer verbatim. Numbers excluded.
    """
    items = [e for e in entries
             if answer_class(e["answer"]) == "text" and _num(e["answer"]) is None]
    if len(items) < MIN_PEERS:
        return None
    groups: list[list[dict]] = []
    for e in items:
        g = next((g for g in groups if _answer_sim(e["answer"], g[0]["answer"], avecs) >= cos_thr), None)
        (g.append(e) if g is not None else groups.append([e]))
    groups.sort(key=len, reverse=True)
    shown = []
    for g in groups:
        if len(g) <= 2:  # minorities/singletons shown individually (precise mapping)
            for e in g:
                shown.append({"files": [e["file"]], "text": e["answer"], "count": 1})
        else:            # collapse a majority cluster to one representative + count
            shown.append({"files": [e["file"] for e in g], "text": g[0]["answer"], "count": len(g)})
    return {"question": title, "items": shown[:max_answers]}


def _parse_flagged(content: str) -> list[dict]:
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        import re
        m = re.search(r"\{.*\}", content, re.S)
        data = json.loads(m.group(0)) if m else {}
    return data.get("flagged", []) if isinstance(data, dict) else []


def _llm_client(cfg: dict):
    from langchain_openai import AzureChatOpenAI
    az = cfg["azure"]
    key = os.environ.get("AZURE_OPENAI_API_KEY") or az.get("api_key")
    if not key:
        raise RuntimeError("no Azure key (set AZURE_OPENAI_API_KEY or azure.api_key)")
    dep = cfg.get("llm_review", {}).get("deployment") or az["deployment"]
    return AzureChatOpenAI(azure_endpoint=az["endpoint"], azure_deployment=dep,
                           api_version=az["api_version"], api_key=key, temperature=0)


def _judge_slot(client, payload: dict) -> list[dict]:
    lines = [f'{i}: {it["text"][:400]}' + (f'  (given by {it["count"]} firms)' if it["count"] > 1 else "")
             for i, it in enumerate(payload["items"])]
    human = f'Question: {payload["question"]}\nPeer answers:\n' + "\n".join(lines)
    resp = client.invoke([("system", LLM_SYSTEM), ("human", human)])
    return _parse_flagged(resp.content)


def llm_layer(report: dict, cfg: dict | None, cache_path: Path, avecs, cos_thr: float,
              max_answers: int = 40, judge=None) -> dict:
    """Per-question LLM anomaly pass over the already-clustered report.

    ``judge`` (payload -> flagged list) is injectable so tests run offline; when
    None a real Azure client is built lazily and only for cache misses.
    Returns {cells, findings} parallel to the rule-based severity grid.
    """
    fnames = report["files"]
    clusters = report["clusters"]
    cells = {fn: [-1] * len(clusters) for fn in fnames}  # -1 = absent / not judged
    cache: dict = {}
    if cache_path.exists():
        cache = json.loads(cache_path.read_text(encoding="utf-8"))

    tasks = []  # (ci, payload, key)
    for ci, c in enumerate(clusters):
        entries = c["slots"][0]["entries"] if c["slots"] else []
        for e in entries:  # baseline 0 for every substantive text answer present
            if e["file"] in cells and answer_class(e["answer"]) == "text" and _num(e["answer"]) is None:
                cells[e["file"]][ci] = 0
        payload = _compress_slot(c["title"], entries, avecs, cos_thr, max_answers)
        if payload is None:
            continue
        key = hashlib.sha256((LLM_PROMPT_VERSION + "\x00" +
                              json.dumps(payload, sort_keys=True, ensure_ascii=False)
                              ).encode("utf-8")).hexdigest()
        tasks.append((ci, payload, key))

    missing = [t for t in tasks if t[2] not in cache]
    if missing:
        judge_fn = judge
        if judge_fn is None:
            client = _llm_client(cfg)
            judge_fn = lambda p: _judge_slot(client, p)  # noqa: E731
        with ThreadPoolExecutor(max_workers=8) as ex:
            for (ci, payload, key), flagged in zip(missing, ex.map(lambda t: judge_fn(t[1]), missing)):
                cache[key] = flagged
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(json.dumps(cache), encoding="utf-8")

    findings = []
    for ci, payload, key in tasks:
        for fl in cache.get(key, []):
            idx, kind = fl.get("id"), fl.get("kind", "outlier")
            if not isinstance(idx, int) or not (0 <= idx < len(payload["items"])):
                continue
            sev = KIND_SEV.get(kind, 2)
            item = payload["items"][idx]
            for fn in item["files"]:
                if fn in cells:
                    cells[fn][ci] = max(cells[fn][ci], sev)
                findings.append({"file": fn, "cluster": ci, "kind": kind, "severity": sev,
                                 "rationale": fl.get("rationale", ""), "answer": item["text"]})
    findings.sort(key=lambda g: (-g["severity"], g["cluster"], g["file"]))
    return {"cells": cells, "findings": findings}


def render_llm_section(report: dict, llm: dict) -> str:
    """HTML fragment (heat map + anomaly detail) spliced into the base report."""
    files = report["files"]
    ctitles = [c["title"] for c in report["clusters"]]
    esc = lambda s: html.escape(str(s))  # noqa: E731
    rat = {(f["file"], f["cluster"]): f for f in llm["findings"]}

    z, t = [], []
    for fn in files:
        zr, tr = [], []
        for ci, sev in enumerate(llm["cells"][fn]):
            if sev < 0:
                zr.append(None); tr.append("")
            else:
                zr.append(sev)
                fnd = rat.get((fn, ci))
                tr.append(f"<b>{LLM_NAMES[sev]}</b>" + (f"<br>{esc(fnd['rationale'][:160])}" if fnd else ""))
        z.append(zr); t.append(tr)
    fig = kyd._heatmap(z, kyd._uniq_labels(ctitles, 34), files, t, LLM_COLORS,
                       "canonical question", "questionnaire",
                       "%{y}<br>%{x}<br>%{text}<extra></extra>")
    plot = fig.to_html(full_html=False, include_plotlyjs=False,
                       config={"displaylogo": False, "responsive": True})
    legend = kyd._legend(LLM_COLORS, LLM_NAMES, "no substantive answer / question absent")

    items = "".join(
        f'<li><span class="chip l{4 if f["severity"] >= 3 else f["severity"]}">{esc(f["kind"])}</span> '
        f'<b>{esc(f["file"])}</b> × {esc(ctitles[f["cluster"]][:60])} — {esc(f["rationale"])} '
        f'<span class="muted">“{esc(f["answer"][:120])}”</span></li>'
        for f in llm["findings"])
    detail = f'<ol id="findings">{items}</ol>' if items else '<p class="muted">No anomalies flagged.</p>'

    return (
        '<h2>LLM anomaly review (supplement)</h2>'
        f'<p class="sub">Peer-context judgement, no domain rules — {len(llm["findings"])} answer(s) flagged. '
        'Rows = questionnaires, columns = questions. Hover a cell for the reason.</p>'
        f'{legend}<div class="plot">{plot}</div>'
        '<h2>LLM anomaly detail</h2>'
        f'{detail}')


# --- output ------------------------------------------------------------------

def write_xlsx(report: dict, stats: list[dict], llm: dict | None, path: Path) -> None:
    """kyd_review's sheets, plus a Summary sheet inserted as the first tab."""
    render_xlsx(report, path)  # Heatmap / wording / Findings / Questions / Answers
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb.create_sheet("Summary", 0)
    headers = ["Questionnaire", "Sections", "Questions", "Answered",
               "Missing (blank)", "Missing %", "N/A", "N/A %"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for s in stats:
        ws.append([s["name"], s["sections"], s["questions"], s["answered"],
                   s["missing"], s["missing_pct"], s["na"], s["na_pct"]])
    for i, w in enumerate([30, 10, 11, 11, 15, 11, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    if llm and llm["findings"]:  # supplementary LLM anomalies, as a second tab
        titles = [c["title"] for c in report["clusters"]]
        wl = wb.create_sheet("LLM anomalies", 1)
        wl.append(["File", "Question", "Kind", "Severity", "Rationale", "Answer"])
        for c in wl[1]:
            c.font = Font(bold=True)
        for f in llm["findings"]:
            wl.append([f["file"], titles[f["cluster"]], f["kind"], LLM_NAMES[f["severity"]],
                       f["rationale"], f["answer"]])
        for i, w in enumerate([22, 45, 14, 24, 60, 70], start=1):
            wl.column_dimensions[get_column_letter(i)].width = w
        wl.freeze_panes = "A2"

    wb.save(path)


# --- cli ---------------------------------------------------------------------

def _glob_xlsx(d: Path) -> list[Path]:
    try:
        files = sorted(d.glob("*.xlsx"))
    except OSError:
        files = sorted(_winlong(d).glob("*.xlsx"))
    return [f for f in files if not f.name.startswith("~$")]  # skip Excel lock files


def main() -> None:
    import sys
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("xlsx", nargs="*", help="long-format questionnaire .xlsx files")
    ap.add_argument("--dir", action="append", metavar="DIR",
                    help="review every *.xlsx in this directory; repeat to merge folders")
    ap.add_argument("-o", "--out", default="output/questionnaire_review.html")
    ap.add_argument("--config", default="config.yaml")
    ap.add_argument("--embed-cache", default=".q_embed_cache.json")
    ap.add_argument("--llm-cache", default=".q_llm_cache.json")
    ap.add_argument("--no-embed", action="store_true",
                    help="skip embeddings; match questions and group answers by difflib only")
    ap.add_argument("--no-llm", action="store_true",
                    help="skip the supplementary LLM anomaly pass")
    args = ap.parse_args()

    dirs = args.dir or []
    multi_folder = len(dirs) > 1
    entries = [(Path(p), None) for p in args.xlsx]
    for d in dirs:
        dp = Path(d)
        stem = dp.resolve().name
        entries += [(f, stem if multi_folder else None) for f in _glob_xlsx(dp)]
    if len(entries) < 2:
        ap.error("need at least 2 xlsx files (pass paths or --dir, repeatable)")

    files = load_files(entries)
    stats = [file_stats(f) for f in files]

    cfg = None
    try:
        cfg = load_config(args.config)
    except Exception as e:  # noqa: BLE001
        print(f"config unavailable ({e}); embeddings and LLM disabled")

    qvecs = avecs = threshold = None
    cos_thr = 0.84
    if cfg and not args.no_embed:
        try:
            threshold = cfg.get("drift", {}).get("cluster_threshold")
            cos_thr = cfg.get("outliers", {}).get("answer_threshold", cos_thr)
            qsigs = {norm(q_signature(q)) for f in files for q in f["questions"]}
            # embed only real free text — pure numbers are judged numerically, not embedded
            atexts = {norm(q["answer"]) for f in files for q in f["questions"]
                      if answer_class(q["answer"]) == "text" and _num(q["answer"]) is None}
            allvecs = embed_signatures(sorted(qsigs | atexts), cfg, Path(args.embed_cache))
            qvecs = {s: allvecs[s] for s in qsigs}
            avecs = {s: allvecs[s] for s in atexts}
            print(f"embeddings: {len(qsigs)} questions, {len(atexts)} free-text answers")
        except Exception as e:  # noqa: BLE001 — any embedding failure -> lexical fallback
            print(f"embeddings unavailable ({e}); falling back to difflib")
            qvecs = avecs = None
            cos_thr = 0.75

    report = analyze(files, multi_folder, qvecs, avecs, threshold, cos_thr)

    llm = None
    if cfg and not args.no_llm:
        try:
            max_answers = cfg.get("llm_review", {}).get("max_answers_per_slot", 40)
            llm = llm_layer(report, cfg, Path(args.llm_cache), avecs, cos_thr, max_answers)
            print(f"LLM anomaly review: {len(llm['findings'])} answer(s) flagged")
        except Exception as e:  # noqa: BLE001 — LLM is a supplement; never fail the run
            print(f"LLM review skipped ({e})")
            llm = None

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    page = render_html(report)
    if llm:
        page = page.replace("</body>", render_llm_section(report, llm) + "</body>", 1)
    out.write_text(page, encoding="utf-8")
    xlsx = out.with_suffix(".xlsx")
    write_xlsx(report, stats, llm, xlsx)

    print(f"\n{len(files)} questionnaires, {len(report['clusters'])} canonical questions")
    print(f"{'questionnaire':<34}{'sect':>5}{'q':>4}{'miss%':>7}{'na%':>6}")
    for s in stats:
        print(f"{s['name'][:33]:<34}{s['sections']:>5}{s['questions']:>4}"
              f"{s['missing_pct']:>7}{s['na_pct']:>6}")
    counts = Counter(g["level"] for g in report["findings"])
    for lv in (OUTLIER, MISSING, WARN, NOTE):
        print(f"  {kyd.LEVEL_NAMES[lv]:>18}: {counts.get(lv, 0)}")
    print(f"report -> {out} and {xlsx}")


def _selfcheck() -> None:
    """Numeric outlier + free-text outlier must fire; pure numbers never embedded."""
    def q(qid, question, section, answer):
        return {"question_id": qid, "question": question, "section": section,
                "answer": answer, "sub_questions": []}

    turnovers = ["15000000", "16000000", "14500000", "15500000", "15200000", "99000000"]  # last = outlier
    sofs = ["freight service revenues"] * 5 + ["proceeds from unrelated crypto trading"]  # last diverges
    files = []
    for i in range(6):
        files.append({"name": f"f{i}", "folder": "", "raw_name": f"f{i}", "questions": [
            q("Q1", "What is the expected annual turnover?", "4. Financial", turnovers[i]),
            q("Q2", "What is the primary source of funds?", "4. Financial", sofs[i]),
        ]})
    rep = analyze(files, False, None, None, None, 0.75)
    kinds = {f["kind"] for f in rep["findings"]}
    assert "numeric outlier" in kinds, kinds
    assert "free-text outlier" in kinds, kinds
    # pure numbers must be excluded from the embed set
    atexts = {norm(qq["answer"]) for f in files for qq in f["questions"]
              if answer_class(qq["answer"]) == "text" and _num(qq["answer"]) is None}
    assert all(_num(a) is None for a in atexts), "numbers leaked into embed set"
    assert file_stats(files[0]) == {"name": "f0", "sections": 1, "questions": 2, "answered": 2,
                                    "missing": 0, "missing_pct": 0.0, "na": 0, "na_pct": 0.0}

    # LLM layer, offline: stub judge flags any shown item mentioning "crypto".
    # Verifies (a) the flag maps to the right file/cell, (b) pure-number questions
    # are never sent to the judge (Q1 must produce no payload, no findings).
    seen_questions = []

    def stub_judge(payload):
        seen_questions.append(payload["question"])
        return [{"id": i, "kind": "implausible", "rationale": "unrelated source"}
                for i, it in enumerate(payload["items"]) if "crypto" in it["text"].lower()]

    cache = Path("__selfcheck_llm_cache.json")
    try:
        llm = llm_layer(rep, None, cache, None, 0.75, judge=stub_judge)
    finally:
        cache.unlink(missing_ok=True)
    assert any("turnover" not in qn for qn in seen_questions)  # source-of-funds judged
    assert all("turnover" not in qn for qn in seen_questions), "numbers reached the LLM"
    hit = [f for f in llm["findings"] if "crypto" in f["answer"].lower()]
    assert hit and hit[0]["file"] == "f5" and hit[0]["severity"] == 3, llm["findings"]
    q2 = next(i for i, c in enumerate(rep["clusters"]) if "source of funds" in c["title"].lower())
    assert llm["cells"]["f5"][q2] == 3
    assert render_llm_section(rep, llm).count("crypto") >= 1
    print("selfcheck ok")


if __name__ == "__main__":
    import sys
    if "--selfcheck" in sys.argv:
        _selfcheck()
    else:
        main()
