"""
Demo questionnaire generator
============================
Creates a sample .xlsx that mirrors the structures seen in the real KYD
questionnaire screenshots, so you can validate the parser against known input.

Includes:
  • Section headers (bold, large, uppercase, centred)
  • Q-numbers in col L, question text merged across N:AC
  • Grey-filled answer slots (some empty, some populated)
  • One yellow-filled answer (for fill-variant testing)
  • One green-highlighted question + a green-highlighted option label
  • Italic hint / sub-label text
  • Dropdown side-answers (YES/NO etc.) in merged B:G
  • Q19-style multi-branch question (multiple side answers per Q)
  • Two stray cells OUTSIDE B:AO to exercise the remarks capture

Usage:
    pip install openpyxl
    python generate_demo.py [output_path]

Default output: ./demo_questionnaire.xlsx
"""

from __future__ import annotations
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

# ── Layout ──────────────────────────────────────────────────────────────────
SIDE_START, SIDE_END     = "B", "G"     # side-answer dropdown lives here
QNUM_COL                 = "L"
QTEXT_START, QTEXT_END   = "N", "AC"    # question/answer text spans here

def cidx(letter: str) -> int:
    return column_index_from_string(letter)

# ── Styles ──────────────────────────────────────────────────────────────────
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")

HEADER_FONT = Font(bold=True, size=14)
ITALIC_FONT = Font(italic=True, color="595959")

THIN = Side(border_style="thin", color="888888")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

# ── Builder helpers ─────────────────────────────────────────────────────────
def merge_set(ws, row, col_start, col_end, value="", fill=None, font=None,
              align=None, border=None, height=None):
    """Merge a horizontal range and set the anchor cell's value/style."""
    ws.merge_cells(start_row=row, end_row=row,
                   start_column=cidx(col_start), end_column=cidx(col_end))
    cell = ws.cell(row=row, column=cidx(col_start))
    cell.value = value
    if fill:   cell.fill = fill
    if font:   cell.font = font
    if align:  cell.alignment = align
    if border: cell.border = border
    if height: ws.row_dimensions[row].height = height
    return cell

def section_header(ws, row, text):
    merge_set(ws, row, QTEXT_START, QTEXT_END,
              value=text, font=HEADER_FONT, align=CENTER, border=BOX, height=26)
    return row + 2

def question_row(ws, row, qnum, qtext, fill=None):
    ws.cell(row=row, column=cidx(QNUM_COL), value=qnum)
    merge_set(ws, row, QTEXT_START, QTEXT_END,
              value=qtext, fill=fill, align=WRAP, height=22)
    return row + 1

def italic_label(ws, row, text):
    merge_set(ws, row, QTEXT_START, QTEXT_END,
              value=text, font=ITALIC_FONT, align=WRAP)
    return row + 1

def plain_text(ws, row, text, fill=None):
    merge_set(ws, row, QTEXT_START, QTEXT_END,
              value=text, fill=fill, align=WRAP)
    return row + 1

def answer_box(ws, row, text="", fill=None):
    merge_set(ws, row, QTEXT_START, QTEXT_END,
              value=text, fill=fill or GREY_FILL,
              align=WRAP, border=BOX, height=32)
    return row + 1

def side_dropdown(ws, row, value, options="YES,NO"):
    """Add a merged cell with a list data validation in the side area."""
    merge_set(ws, row, SIDE_START, SIDE_END,
              value=value, fill=GREY_FILL, border=BOX, align=WRAP)
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{SIDE_START}{row}")

# ── Main ────────────────────────────────────────────────────────────────────
def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questionnaire HK"

    # Column widths (rough match to the screenshots)
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFG":   ws.column_dimensions[col].width = 4
    for col in "HIJK":     ws.column_dimensions[col].width = 2
    ws.column_dimensions["L"].width = 6
    ws.column_dimensions["M"].width = 2
    for col in list("NOPQRSTUVWXYZ") + ["AA", "AB", "AC"]:
        ws.column_dimensions[col].width = 5

    row = 2

    # ─── Section 1: ORGANISATION ─────────────────────────────────────────
    row = section_header(ws, row, "ORGANISATION, BUSINESS DESCRIPTION & EXPERIENCE")

    # Q1 - simple Q&A
    row = question_row(ws, row, "Q1", "Firm Full Legal Name and Address")
    row = answer_box (ws, row, "Acme Bank (Hong Kong) Limited")
    row += 1

    # Q2 - with italic hint
    row = question_row(ws, row, "Q2", "Regulatory Authority")
    row = italic_label(ws, row, "SFC, Monetary Authority of Singapore etc.")
    row = answer_box  (ws, row, "SFC & HKMA")
    row += 1

    # Q3 - labelled-section question: a short title stem with two labelled
    # sections. HK is filled with a multi-line list answer; Singapore shows only
    # a placeholder hint and is left blank. A large gap separates the sections.
    row = question_row(ws, row, "Q3", "Relevant Authorised Activities")
    row = italic_label(ws, row, "For HK, please specify license type")
    row = answer_box  (ws, row,
        "Authorized Institution - HKMA\n"
        "Dealing in Securities - SFC\n"
        "Advising on Securities - SFC\n"
        "Advising on Securities - SFC")
    row += 6  # large vertical gap between sections
    row = italic_label(ws, row, "For Singapore, please specify license type")
    row = italic_label(ws, row, "Please detail the license type and conditions.")
    row = answer_box  (ws, row, "")  # placeholder only — left blank
    row += 1

    # Q4 - hint + answer
    row = question_row(ws, row, "Q4", "Economic data (If available: number of staff, AUM, ...)")
    row = answer_box  (ws, row, "Approx no of employees: 500, AUM: HK$ eqv 80 billion")
    row += 1

    # ─── Section 2: PRODUCT APPROVAL PROCESS ─────────────────────────────
    row = section_header(ws, row, "PRODUCT APPROVAL PROCESS")

    # Q5 - GREEN-highlighted question + YES side-answer (like Q14 in screenshots)
    side_dropdown(ws, row, "YES")
    row = question_row(ws, row, "Q5",
        "a) Has your firm adopted a product approval process with respect to structured products?",
        fill=GREEN_FILL)
    row = answer_box(ws, row, "Yes")
    row = plain_text(ws, row, "b) Does each structured product go through the same approval process?")
    row = answer_box(ws, row,
        "Yes, each structured product type will go through the same approval process.")
    row += 1

    # Q6 - YELLOW-highlighted answer (variant fill)
    row = question_row(ws, row, "Q6",
        "What are the key features reviewed in the product approval process for structured products?")
    row = italic_label(ws, row,
        "This includes but not limited to: Product risk classification, "
        "investor categorization, risk mapping to investor and product etc")
    row = answer_box(ws, row,
        "Product Control Committee reviews all new structured products based on "
        "risk classification, target customer profile, and cost-benefit analysis.",
        fill=YELLOW_FILL)
    row += 1

    # ─── Section 3: DISTRIBUTION CHANNEL (Q19-style multi-branch) ────────
    row = section_header(ws, row, "DISTRIBUTION CHANNEL")

    row = question_row(ws, row, "Q7",
        "In which capacity will your firm act as distributor?")

    # Branch 1: "Settling distributor" (green) + YES side + empty grey answer
    side_dropdown(ws, row, "YES")
    row = plain_text(ws, row, "Settling distributor", fill=GREEN_FILL)
    row = plain_text(ws, row, "Who is the custodian?")
    row = answer_box(ws, row, "")  # empty answer slot

    # Branch 2: "Non-settling distributor" + NO side + answer "PERSHING LLC"
    side_dropdown(ws, row, "NO")
    row = plain_text(ws, row, "Non-settling distributor")
    row = plain_text(ws, row, "Who is the custodian?")
    row = answer_box(ws, row, "PERSHING LLC")

    # Branch 3: PtP question with empty answer
    row = plain_text(ws, row, "Is this PtP set up? i.e you act as principal rather than agent.")
    row = answer_box(ws, row, "")
    row += 1

    # ─── Stray cells OUTSIDE B:AO (should land in `remarks`) ─────────────
    ws.cell(row=10, column=cidx("AP"),
            value="External reviewer note: please verify Q1 address.").font = ITALIC_FONT
    ws.cell(row=25, column=cidx("AQ"),
            value="Internal flag: Q6 answer needs follow-up.").font = ITALIC_FONT

    return wb

def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("./demo_questionnaire.xlsx")
    wb = build_workbook()
    wb.save(out)
    print(f"✓ Demo workbook written to: {out.resolve()}")
    print(f"\nTry it with the parser:")
    print(f"  python parse_questionnaire.py {out} ./output --debug")

if __name__ == "__main__":
    main()