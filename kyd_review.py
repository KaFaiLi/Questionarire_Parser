"""Cross-questionnaire review of parsed KYD JSON files.

Reads every parsed questionnaire (the ``questions[] -> sub_questions[]`` format
produced by parse_questionnaires_llm.py), matches the *same* question across
files by text similarity (question ids are NOT trusted — the same question can
be Q10 in one file and Q12 in another), then checks:

Questions
  - wording drift per canonical question: formatting-only vs substantive
  - questions present in only a few files (questionnaire-specific)
  - common questions missing from a file

Answers (per canonical question, per aligned sub-question slot)
  - required answers left blank (blank-in-a-minority => missing;
    blank-in-the-majority => conditional branch, not a defect)
  - categorical outliers (e.g. 9x YES / 1x NO)
  - numeric outliers (robust median/MAD)
  - free-text minority answers diverging from the common wording
  - YES/NO selection contradicting the written answer

Output: one self-contained HTML report (interactive Plotly heat maps of files x
questions, hover for detail) plus a matching xlsx and a console summary. The HTML
renders with no server and no external requests — only an inline Plotly bundle.

    python kyd_review.py --dir Demo/kyd_examples -o output/kyd_review.html
    python kyd_review.py --dir //server/share/batch1 --dir //server/share/batch2 -o out.html

Questions are matched semantically via Azure OpenAI embeddings (cached on disk so
re-runs are free); with --no-embed, or when no Azure key is configured, it falls
back to a stdlib difflib text scan.
"""
from __future__ import annotations

import argparse
import difflib
import hashlib
import html
import json
import os
import re
import statistics
from collections import Counter
from pathlib import Path

# severity levels
OK, NOTE, WARN, MISSING, OUTLIER = 0, 1, 2, 3, 4
LEVEL_NAMES = {OK: "ok", NOTE: "formatting", WARN: "wording / specific",
               MISSING: "missing", OUTLIER: "answer outlier"}

SIM_QUESTION = 0.60   # min similarity to treat two questions as the same
SIM_SLOT = 0.70       # min similarity to align two sub-question slots
SIM_TEXT_GROUP = 0.75 # min similarity to group two free-text answers

# Question matching profiles. Tolerant mode is for parser/OCR-heavy inputs. It
# keeps the full q_signature() for initial clustering, including its
# sub-question fallback for records with no top-level question text, then uses
# high-confidence parent evidence to reconnect parser fragments.
QUESTION_MATCHING = {
    "standard": {"embedding_threshold": 0.84, "slot_similarity": SIM_SLOT,
                 "parent_presence": False},
    "tolerant": {"embedding_threshold": 0.78, "slot_similarity": 0.60,
                 "parent_presence": True},
}
PARENT_PRESENCE_DEFAULTS = {
    "allow_top_level_evidence": True,
    "top_level_similarity": 0.78,
    "allow_subquestion_evidence": False,
    "slot_similarity": 0.78,
    "minimum_distinctive_slot_matches": 1,
    "uniqueness_margin": 0.05,
    "ignore_subquestion_wording_drift": False,
}
MINORITY_SHARE = 0.25 # answer value held by <= this share => outlier candidate

# Focused-review defaults.  These deliberately require broad peer agreement
# before turning an inconsistency into a finding.  They can be overridden by
# the ``review`` section in config.yaml.
SPECIFIC_SHARE = 0.15           # <= this share: questionnaire-specific (informational)
COMMON_QUESTION_SHARE = 0.85    # >= this share: absence is a missing question
WORDING_MIN_COVERAGE = 0.80     # wording differences need this question coverage
WORDING_VARIANT_MAX_SHARE = 0.25  # and must be a small minority variant
REQUIRED_SHARE = 0.85           # peers must answer at this rate before blank is missing
NA_SUBSTANTIVE_SHARE = 0.85     # peers must give text at this rate before N/A is challenged


def norm(s: str) -> str:
    """Case/whitespace/punctuation-insensitive form for comparisons."""
    return re.sub(r"[^a-z0-9 ]+", "", re.sub(r"\s+", " ", (s or "").lower())).strip()


# "N/A", "n.a.", "NA", "Not applicable", "Nil", "None" — a declared non-answer,
# distinct from a blank: the firm asserts the question does not apply.
NA_VALUES = {"na", "n a", "not applicable", "nil", "none"}
YESNO = {"y": "YES", "yes": "YES", "n": "NO", "no": "NO"}


def answer_class(s: str) -> str:
    """'blank' | 'na' (declared not-applicable) | 'text' (substantive)."""
    t = norm(s)
    return "blank" if not t else ("na" if t in NA_VALUES else "text")


def canon(s: str) -> str:
    """Comparison value: yes/no synonyms collapsed (Y == yes == YES)."""
    t = norm(s)
    return YESNO.get(t, t)


def ratio(a: str, b: str) -> float:
    return difflib.SequenceMatcher(None, a, b).ratio()


# --- embeddings (semantic question matching, with an on-disk cache) ----------

def load_config(path: str = "config.yaml") -> dict:
    import yaml
    return yaml.safe_load(Path(path).read_text(encoding="utf-8"))


def question_matching_settings(cfg: dict) -> dict:
    """Resolve the question matching profile and validate its overrides.

    ``drift.cluster_threshold`` remains the standard-mode compatibility
    setting.  In tolerant mode, ``question_matching.embedding_threshold`` and
    ``question_matching.slot_similarity`` may override the profile defaults.
    """
    settings = cfg.get("question_matching", {}) or {}
    mode = settings.get("mode", "standard")
    if mode not in QUESTION_MATCHING:
        choices = ", ".join(sorted(QUESTION_MATCHING))
        raise ValueError(f"question_matching.mode must be one of: {choices}")
    defaults = QUESTION_MATCHING[mode]

    def threshold(name: str, default: float) -> float:
        value = float(settings.get(name, default))
        if not 0 < value <= 1:
            raise ValueError(f"question_matching.{name} must be greater than 0 and at most 1")
        return value

    standard_embedding = cfg.get("drift", {}).get("cluster_threshold", defaults["embedding_threshold"])
    embedding_default = standard_embedding if mode == "standard" else defaults["embedding_threshold"]
    raw_parent = settings.get("parent_presence", {}) or {}
    if not isinstance(raw_parent, dict):
        raise ValueError("question_matching.parent_presence must be a mapping")

    def flag(name: str, default: bool) -> bool:
        value = raw_parent.get(name, default)
        if not isinstance(value, bool):
            raise ValueError(f"question_matching.parent_presence.{name} must be true or false")
        return value

    def parent_threshold(name: str, default: float) -> float:
        value = float(raw_parent.get(name, default))
        if not 0 < value <= 1:
            raise ValueError(f"question_matching.parent_presence.{name} must be greater than 0 and at most 1")
        return value

    enabled = flag("allow_subquestion_evidence", defaults["parent_presence"])
    parent_presence = {
        "allow_top_level_evidence": flag("allow_top_level_evidence", True),
        "top_level_similarity": parent_threshold("top_level_similarity", PARENT_PRESENCE_DEFAULTS["top_level_similarity"]),
        "allow_subquestion_evidence": enabled,
        # This is deliberately higher than tolerant slot alignment: a slot can
        # be loosely aligned for answer review but must be a strong signal to
        # prove that a parent question exists.
        "slot_similarity": parent_threshold("slot_similarity", PARENT_PRESENCE_DEFAULTS["slot_similarity"]),
        "minimum_distinctive_slot_matches": int(raw_parent.get(
            "minimum_distinctive_slot_matches", PARENT_PRESENCE_DEFAULTS["minimum_distinctive_slot_matches"])),
        "uniqueness_margin": parent_threshold("uniqueness_margin", PARENT_PRESENCE_DEFAULTS["uniqueness_margin"]),
        "ignore_subquestion_wording_drift": flag("ignore_subquestion_wording_drift", enabled),
    }
    if parent_presence["minimum_distinctive_slot_matches"] < 1:
        raise ValueError("question_matching.parent_presence.minimum_distinctive_slot_matches must be at least 1")

    return {"mode": mode,
            "embedding_threshold": threshold("embedding_threshold", embedding_default),
            "slot_similarity": threshold("slot_similarity", defaults["slot_similarity"]),
            "parent_presence": parent_presence}


def _azure_key(az: dict) -> str | None:
    return os.environ.get("AZURE_OPENAI_API_KEY") or az.get("api_key")


def embed_signatures(sigs: list[str], cfg: dict, cache_path: Path) -> dict[str, list[float]]:
    """Map each unique signature -> embedding vector.

    Embeds only signatures not already in the JSON cache, so re-runs on the same
    corpus cost nothing. Raises on Azure/import failure; callers fall back to difflib.
    """
    az = cfg["azure"]
    key = _azure_key(az)
    if not key:
        raise RuntimeError("no Azure key (set AZURE_OPENAI_API_KEY or azure.api_key)")

    cache: dict[str, list[float]] = {}
    if cache_path.exists():
        cache = json.loads(cache_path.read_text(encoding="utf-8"))

    uniq = sorted(set(sigs))
    keys = {s: hashlib.sha256(s.encode("utf-8")).hexdigest() for s in uniq}
    missing = [s for s in uniq if keys[s] not in cache]
    if missing:
        from langchain_openai import AzureOpenAIEmbeddings
        emb = AzureOpenAIEmbeddings(
            azure_endpoint=az["endpoint"],
            azure_deployment=az["embedding_deployment"],
            api_version=az["api_version"],
            api_key=key,
        )
        for vec, s in zip(emb.embed_documents(missing), missing):
            cache[keys[s]] = vec
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(json.dumps(cache), encoding="utf-8")

    return {s: cache[keys[s]] for s in uniq}


def q_signature(q: dict) -> str:
    """Text that identifies a question: heading + prompts + option labels."""
    parts = [q.get("question", "")]
    for s in q.get("sub_questions", []):
        parts += [s.get("option_label", ""), s.get("prompt", "")]
    return " ".join(p for p in parts if p)


def _winlong(p: Path) -> Path:
    """Extended-length form (\\\\?\\...) so paths over MAX_PATH (260 chars) work.

    ponytail: string-prefix trick, not a general long-path fix (relative paths,
    '.'/'..' segments still need resolve() first); good enough for UNC network
    shares with deep folders, which is the actual failure seen.
    """
    if os.name != "nt":
        return p
    s = str(p)
    if s.startswith("\\\\?\\"):
        return p
    rp = str(p.resolve())
    return Path("\\\\?\\UNC\\" + rp[2:]) if rp.startswith("\\\\") else Path("\\\\?\\" + rp)


def _glob_json(d: Path) -> list[Path]:
    try:
        return sorted(d.glob("*.json"))
    except OSError:
        return sorted(_winlong(d).glob("*.json"))


def _read_json(p: Path) -> dict:
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError):
        return json.loads(_winlong(p).read_text(encoding="utf-8"))


def load_files(entries: list[tuple[Path, str | None]]) -> list[dict]:
    """entries: (path, folder_stem_or_None). folder_stem set only when >1 --dir given."""
    raw = []
    for p, folder in sorted(entries, key=lambda e: (e[1] or "", e[0].name)):
        d = _read_json(p)
        raw.append({"raw_name": d.get("file_name") or p.stem, "folder": folder or "",
                    "questions": d.get("questions", [])})
    dupes = Counter(f["raw_name"] for f in raw)
    for f in raw:
        f["name"] = f["raw_name"] if dupes[f["raw_name"]] == 1 else f"{f['raw_name']} ({f['folder']})"
    return raw


# --- question matching -----------------------------------------------------

def cluster_questions(files: list[dict], vecs: dict[str, list[float]] | None = None,
                      threshold: float | None = None) -> list[dict]:
    """Greedy clustering of questions across files.

    With ``vecs`` (signature -> embedding) it matches by cosine similarity, so
    paraphrases group together. Without it, falls back to a difflib character
    scan (lexical only). Both keep one question per file per cluster.

    ponytail: O(files x questions x clusters) scan — fine for hundreds of files;
    add a blocking/ANN prefilter if it ever gets slow.
    """
    embed = vecs is not None
    thr = threshold if threshold is not None else (0.84 if embed else SIM_QUESTION)
    if embed:
        import numpy as np

        def unit(s):
            v = np.asarray(vecs[s], dtype=float)
            n = np.linalg.norm(v)
            return v / n if n else v

    clusters: list[dict] = []
    for f in files:
        for pos, q in enumerate(f["questions"]):
            sig = norm(q_signature(q))
            u = unit(sig) if embed else None
            best, best_r = None, 0.0
            for c in clusters:
                if f["name"] in c["members"]:
                    continue  # one question per file per cluster
                if embed:
                    cen = c["vsum"]
                    r = float(u @ (cen / (np.linalg.norm(cen) or 1)))
                else:
                    r = ratio(sig, c["sig"])
                if r > best_r:
                    best, best_r = c, r
            if best is not None and best_r >= thr:
                best["members"][f["name"]] = q
                best["positions"].append(pos)
                if embed:
                    best["vsum"] = best["vsum"] + u
            else:
                c = {"sig": sig, "members": {f["name"]: q}, "positions": [pos]}
                if embed:
                    c["vsum"] = u.copy()
                clusters.append(c)
    clusters.sort(key=lambda c: statistics.median(c["positions"]))
    return clusters


def cluster_title(c: dict) -> str:
    texts = [q["question"] for q in c["members"].values() if q.get("question")]
    if not texts:
        texts = [s["prompt"] for q in c["members"].values()
                 for s in q.get("sub_questions", []) if s.get("prompt")]
    return Counter(texts).most_common(1)[0][0] if texts else "(untitled)"


_GENERIC_SLOT_WORDS = {
    "a", "an", "and", "answer", "applicable", "details", "for", "if", "information", "n",
    "na", "no", "not", "of", "or", "please", "provide", "response", "the", "to", "yes",
}


def _slot_raw(s: dict) -> str:
    return (s.get("option_label", "") + " " + s.get("prompt", "")).strip()


def _distinctive_text(raw: str, minimum_length: int) -> bool:
    """Whether text has enough non-boilerplate content to identify a parent."""
    key = norm(raw)
    tokens = [t for t in key.split() if t not in _GENERIC_SLOT_WORDS]
    return len(key) >= minimum_length and len(tokens) >= 2


def _distinctive_slot(raw: str) -> bool:
    """Whether a child prompt is sufficiently specific to identify a parent.

    Generic fragments ("yes/no", "please provide details") are useful for
    alignment inside a known question but must never prove a parent question is
    present: they commonly appear throughout a questionnaire.
    """
    return _distinctive_text(raw, 16)


def _best_slot_similarity(raw: str, target: dict) -> float:
    """Best lexical child-slot match, preserving the enumeration guard."""
    key, mark = norm(raw), _enum_marker(raw)
    scores = []
    for q in target["members"].values():
        for s in q.get("sub_questions", []):
            other = _slot_raw(s)
            if mark != _enum_marker(other):
                continue
            scores.append(ratio(key, norm(other)))
    return max(scores, default=0.0)


def _best_top_level_similarity(raw: str, target: dict) -> float:
    key = norm(raw)
    return max((ratio(key, norm(q.get("question", ""))) for q in target["members"].values()
                if norm(q.get("question", ""))), default=0.0)


def _add_unambiguous_evidence(raw: str, source_file: str, source_cluster: dict,
                              clusters: list[dict], threshold: float, margin: float,
                              scorer, evidence: dict[int, list[float]], targets: dict[int, dict]) -> None:
    """Add evidence for a uniquely best parent cluster, if one exists."""
    scores = []
    for target in clusters:
        if (target is source_cluster or source_file in target["members"]
                or len(target["members"]) <= len(source_cluster["members"])):
            continue
        score = scorer(raw, target)
        if score >= threshold:
            scores.append((score, target))
    scores.sort(key=lambda item: item[0], reverse=True)
    if not scores:
        return
    best_score, best_target = scores[0]
    second_score = scores[1][0] if len(scores) > 1 else 0.0
    if second_score and best_score - second_score < margin:
        return
    key = id(best_target)
    evidence.setdefault(key, []).append(best_score)
    targets[key] = best_target


def _parent_target_for_question(source: dict, source_file: str, source_cluster: dict,
                                clusters: list[dict], presence: dict) -> dict | None:
    """Find one unambiguous parent cluster from heading or child-slot evidence."""
    evidence: dict[int, list[float]] = {}
    targets: dict[int, dict] = {}
    title = source.get("question", "")
    if presence["allow_top_level_evidence"] and _distinctive_text(title, 8):
        _add_unambiguous_evidence(title, source_file, source_cluster, clusters,
                                  presence["top_level_similarity"], presence["uniqueness_margin"],
                                  _best_top_level_similarity, evidence, targets)
    title_targets = set(evidence)
    for s in source.get("sub_questions", []):
        raw = _slot_raw(s)
        if not _distinctive_slot(raw):
            continue
        _add_unambiguous_evidence(raw, source_file, source_cluster, clusters,
                                  presence["slot_similarity"], presence["uniqueness_margin"],
                                  _best_slot_similarity, evidence, targets)

    candidates = [(len(scores), sum(scores), targets[key])
                  for key, scores in evidence.items()
                  if (len(scores) >= presence["minimum_distinctive_slot_matches"]
                      or key in title_targets)]
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    # Two targets with the same evidence strength leave parent presence
    # uncertain.  Suppress the merge rather than guessing.
    if len(candidates) > 1 and candidates[0][:2] == candidates[1][:2]:
        return None
    return candidates[0][2]


def merge_parent_evidence_clusters(clusters: list[dict], presence: dict) -> list[dict]:
    """Merge parser-fragmented clusters when all their members prove one parent.

    Initial clustering remains based on the full question/sub-question
    signature. This second pass only joins a smaller cluster to a larger parent
    cluster with strong, unique heading or child-slot evidence; it therefore
    turns partial child parsing into a parent-presence decision rather than a
    missing parent-question finding.
    """
    if not presence["allow_subquestion_evidence"]:
        return clusters
    while True:
        merged = False
        for source in sorted(clusters, key=lambda c: len(c["members"])):
            targets = [_parent_target_for_question(q, fname, source, clusters, presence)
                       for fname, q in source["members"].items()]
            if not targets or any(t is None for t in targets):
                continue
            target = targets[0]
            if any(t is not target for t in targets[1:]):
                continue
            # The target cannot already contain one of the source files; this
            # maintains the one-question-per-file invariant of a cluster.
            if any(fname in target["members"] for fname in source["members"]):
                continue
            target["members"].update(source["members"])
            target["positions"].extend(source["positions"])
            if "vsum" in target and "vsum" in source:
                target["vsum"] = target["vsum"] + source["vsum"]
            clusters.remove(source)
            merged = True
            break
        if not merged:
            break
    clusters.sort(key=lambda c: statistics.median(c["positions"]))
    return clusters


# --- checks ------------------------------------------------------------------

def check_variants(c: dict, findings: list, ci: int, n_files: int,
                   min_coverage: float, max_variant_share: float,
                   top_level_only: bool = False) -> list[dict]:
    """Classify wording variants of one canonical question."""
    by_raw: dict[str, list[str]] = {}
    unavailable = []
    for fname, q in c["members"].items():
        raw = q.get("question", "") if top_level_only else q_signature(q)
        if top_level_only and not norm(raw):
            unavailable.append(fname)
        else:
            by_raw.setdefault(raw, []).append(fname)
    if not by_raw:
        return [{"text": "", "files": list(c["members"]), "kind": "majority"}]
    variants = [{"text": raw, "files": fs} for raw, fs in by_raw.items()]
    if len(variants) == 1:
        variants[0]["kind"] = "majority"
        if unavailable:
            variants.append({"text": "", "files": unavailable, "kind": "unavailable"})
        return variants
    majority = max(variants, key=lambda v: len(v["files"]))
    for v in variants:
        if v is majority:
            v["kind"] = "majority"
        elif norm(v["text"]) == norm(majority["text"]):
            v["kind"] = "formatting"
            for fn in v["files"]:
                findings.append(dict(file=fn, cluster=ci, level=NOTE, kind="formatting variant",
                                     message="Question wording differs only in spacing/case from the common version."))
        else:
            v["kind"] = "substantive"
            # A genuine wording issue needs both a widely-used question and a
            # clearly exceptional variant.  Otherwise it is useful context in
            # the wording sheet, but not a review finding.
            if (len(c["members"]) / n_files >= min_coverage
                    and len(v["files"]) / n_files <= max_variant_share
                    and len(v["files"]) < len(majority["files"])):
                for fn in v["files"]:
                    findings.append(dict(file=fn, cluster=ci, level=WARN, kind="wording variant",
                                         message=f"Question wording differs substantively from the common version: “{v['text'][:120]}”"))
    if unavailable:
        # A blank top-level heading is a parsing limitation, not wording drift.
        variants.append({"text": "", "files": unavailable, "kind": "unavailable"})
    return variants


def _enum_marker(s: str) -> str:
    """Leading enumeration marker like 'a)', '(ii)', '3.' — '' if none."""
    m = re.match(r"\(?([a-z0-9]{1,3})[\).]", s.strip().lower())
    return m.group(1) if m else ""


def align_slots(c: dict, threshold: float = SIM_SLOT) -> list[dict]:
    """Align sub-questions across files by option label + prompt similarity.

    Slots with different enumeration markers (a) vs b)) are never merged —
    similar wording there means sibling sub-questions, not the same one.
    """
    slots: list[dict] = []
    for fname, q in c["members"].items():
        for s in q.get("sub_questions", []):
            raw = (s.get("option_label", "") + " " + s.get("prompt", "")).strip()
            key, mark = norm(raw), _enum_marker(raw)
            hit = next((sl for sl in slots if sl["mark"] == mark
                        and (sl["key"] == key or ratio(key, sl["key"]) >= threshold)), None)
            if hit is None:
                label = (s.get("option_label") or s.get("prompt") or "answer").strip()
                hit = {"key": key, "mark": mark, "label": label, "entries": []}
                slots.append(hit)
            hit["entries"].append({"file": fname, "selection": s.get("selection", ""),
                                   "answer": s.get("answer", ""), "flags": []})
    return slots


def _num(s: str) -> float | None:
    m = re.fullmatch(r"[~\s]*([-+]?[\d,]+(?:\.\d+)?)\s*%?", s.strip())
    return float(m.group(1).replace(",", "")) if m else None


def _flag(entry: dict, findings: list, ci: int, level: int, kind: str, message: str) -> None:
    entry["flags"].append(kind)
    findings.append(dict(file=entry["file"], cluster=ci, level=level, kind=kind, message=message))


def check_answers(slot: dict, findings: list, ci: int,
                  required_share: float, na_substantive_share: float) -> None:
    entries = slot["entries"]
    n = len(entries)
    if n < 3:
        return

    # YES/NO selection vs written answer contradiction
    for e in entries:
        sel = YESNO.get(norm(e["selection"]), "")
        first = YESNO.get(norm(e["answer"]).split(" ")[0] if e["answer"].strip() else "", "")
        if sel and first and first != sel:
            _flag(e, findings, ci, OUTLIER, "selection/answer mismatch",
                  f"Selection is {e['selection']} but the written answer says “{e['answer'][:80]}”.")

    # categorical minority on the selection column (yes/no synonyms collapsed)
    sels = [e for e in entries if e["selection"].strip()]
    if len(sels) >= 5:
        counts = Counter(canon(e["selection"]) for e in sels)
        if len(counts) <= 4:
            top = counts.most_common(1)[0][1]
            for val, cnt in counts.items():
                if cnt / len(sels) <= MINORITY_SHARE and cnt < top:
                    for e in sels:
                        if canon(e["selection"]) == val:
                            _flag(e, findings, ci, OUTLIER, "selection outlier",
                                  f"Selection “{e['selection']}” given by {cnt}/{len(sels)} firms; the rest answered differently.")

    # blanks vs declared N/A vs substantive answers
    answered = [e for e in entries if answer_class(e["answer"]) != "blank"]
    filled = [e for e in entries if answer_class(e["answer"]) == "text"]
    if answered and len(answered) / n >= required_share:
        for e in entries:
            if answer_class(e["answer"]) == "blank" and not e["selection"].strip():
                _flag(e, findings, ci, MISSING, "missing answer",
                      f"Answer left blank while {len(answered)}/{n} firms answered.")
    # declared N/A while peers answered substantively => the assertion needs challenging
    if answered and len(filled) / len(answered) >= na_substantive_share:
        for e in answered:
            if answer_class(e["answer"]) == "na":
                _flag(e, findings, ci, WARN, "declared N/A",
                      f"Firm answered “{e['answer']}” while {len(filled)}/{len(answered)} peers gave a substantive answer — the not-applicable claim should be challenged.")
    if len(filled) < 5:
        return

    nums = [(_num(e["answer"]), e) for e in filled]
    if sum(1 for v, _ in nums if v is not None) / len(filled) >= 0.8:
        vals = [v for v, _ in nums if v is not None]
        med = statistics.median(vals)
        mad = statistics.median([abs(v - med) for v in vals])
        for v, e in nums:
            if v is None:
                continue
            z = 0.6745 * abs(v - med) / mad if mad else (abs(v - med) / (abs(med) or 1))
            if z > 3.5:
                _flag(e, findings, ci, OUTLIER, "numeric outlier",
                      f"Value {e['answer']} is far from the peer median of {med:g}.")
        return

    # free text: greedy similarity groups, minority group => outlier
    groups: list[list[dict]] = []
    for e in filled:
        t = canon(e["answer"])
        g = next((g for g in groups if ratio(t, canon(g[0]["answer"])) >= SIM_TEXT_GROUP), None)
        (g.append(e) if g is not None else groups.append([e]))
    top = max(len(g) for g in groups)
    if top / len(filled) >= 0.6:
        for g in groups:
            if len(g) / len(filled) <= MINORITY_SHARE and len(g) < top:
                for e in g:
                    _flag(e, findings, ci, OUTLIER, "free-text outlier",
                          f"Answer diverges from the wording used by {top}/{len(filled)} firms: “{e['answer'][:100]}”")


# --- analysis ---------------------------------------------------------------

def analyze(files: list[dict], multi_folder: bool = False,
            vecs: dict[str, list[float]] | None = None, threshold: float | None = None,
            review: dict | None = None, slot_threshold: float = SIM_SLOT,
            parent_presence: dict | None = None) -> dict:
    """Analyze questionnaires using optional focused-review share thresholds."""
    review = review or {}
    parent_presence = {**PARENT_PRESENCE_DEFAULTS, **(parent_presence or {})}

    def share(name: str, default: float) -> float:
        value = float(review.get(name, default))
        if not 0 < value <= 1:
            raise ValueError(f"review.{name} must be greater than 0 and at most 1")
        return value

    specific_share = share("specific_question_share", SPECIFIC_SHARE)
    common_question_share = share("common_question_share", COMMON_QUESTION_SHARE)
    wording_min_coverage = share("wording_min_coverage", WORDING_MIN_COVERAGE)
    wording_variant_max_share = share("wording_variant_max_share", WORDING_VARIANT_MAX_SHARE)
    required_share = share("required_answer_share", REQUIRED_SHARE)
    na_substantive_share = share("na_substantive_share", NA_SUBSTANTIVE_SHARE)
    n_files = len(files)
    clusters = cluster_questions(files, vecs, threshold)
    clusters = merge_parent_evidence_clusters(clusters, parent_presence)
    findings: list[dict] = []
    out_clusters = []
    for ci, c in enumerate(clusters):
        coverage = len(c["members"])
        coverage_share = coverage / n_files
        specific = coverage_share <= specific_share
        expected = coverage_share >= common_question_share
        if specific:
            for fn in c["members"]:
                findings.append(dict(file=fn, cluster=ci, level=NOTE, kind="questionnaire-specific",
                                     message=f"Question appears in only {coverage}/{n_files} questionnaires."))
        elif expected:
            for f in files:
                if f["name"] not in c["members"]:
                    findings.append(dict(file=f["name"], cluster=ci, level=MISSING, kind="question missing",
                                         message=f"Question present in {coverage}/{n_files} questionnaires but absent here."))
        variants = check_variants(c, findings, ci, n_files, wording_min_coverage, wording_variant_max_share,
                                  parent_presence["ignore_subquestion_wording_drift"])
        slots = align_slots(c, slot_threshold)
        for slot in slots:
            check_answers(slot, findings, ci, required_share, na_substantive_share)
        out_clusters.append({
            "title": cluster_title(c),
            "ids": {fn: q["question_id"] for fn, q in c["members"].items()},
            "coverage": coverage, "specific": specific, "expected": expected,
            "variants": variants,
            "slots": [{"label": s["label"], "entries": s["entries"]} for s in slots],
        })

    fnames = [f["name"] for f in files]
    cells = {fn: [] for fn in fnames}
    per_cell = {(g["file"], g["cluster"]): [] for g in findings}
    for g in findings:
        per_cell[(g["file"], g["cluster"])].append(g)
    for fn in fnames:
        for ci, c in enumerate(out_clusters):
            if fn not in c["ids"] and not c["expected"]:
                cells[fn].append(-1)  # not expected here
            else:
                cells[fn].append(max((g["level"] for g in per_cell.get((fn, ci), [])), default=OK))
    findings.sort(key=lambda g: (-g["level"], g["cluster"], g["file"]))
    folders = {f["name"]: f.get("folder", "") for f in files}

    # wording-divergence grid (question x firm): how each firm's wording of a
    # question compares to the majority. 0 same, 1 formatting-only, 2 substantive,
    # -1 absent. Drives the second HTML heat map.
    kindcode = {"majority": 0, "formatting": 1, "substantive": 2}
    qvar = []
    for c in out_clusters:
        fmap = {}
        for v in c["variants"]:
            for fn in v["files"]:
                fmap[fn] = kindcode.get(v.get("kind"), 0)
        qvar.append([fmap.get(fn, -1) for fn in fnames])

    return {"files": fnames, "clusters": out_clusters, "cells": cells, "findings": findings,
            "folders": folders, "multi_folder": multi_folder, "qvar": qvar}


# --- report -----------------------------------------------------------------

# severity levels 0-4 and the wording-divergence codes 0/1/2 each get a fixed colour.
SEV_COLORS = ["#5cb85c", "#9ec5f4", "#fab219", "#ec835a", "#d03b3b"]
SEV_NAMES = {0: "no issue", 1: "formatting-only wording",
             2: "substantive / questionnaire-specific",
             3: "missing (question or required answer)", 4: "answer outlier"}
WORD_COLORS = ["#5cb85c", "#fab219", "#d03b3b"]
WORD_NAMES = {0: "same as majority", 1: "formatting-only difference",
              2: "substantially different wording"}

PAGE_CSS = """
:root { color-scheme: light dark; }
body { margin:0; padding:24px; background:#f9f9f7; color:#0b0b0b;
       font:14px/1.45 system-ui,-apple-system,"Segoe UI",sans-serif; }
h1 { font-size:20px; margin:0 0 4px; } h2 { font-size:16px; margin:28px 0 8px; }
.sub { color:#52514e; margin:0 0 16px; }
.tiles { display:flex; gap:12px; flex-wrap:wrap; margin:16px 0; }
.tile { background:#fff; border:1px solid rgba(11,11,11,.10); border-radius:8px; padding:10px 16px; min-width:110px; }
.tile b { display:block; font-size:24px; } .tile span { color:#52514e; font-size:12px; }
.plot { background:#fff; border:1px solid rgba(11,11,11,.10); border-radius:8px; padding:8px; margin-bottom:8px; }
.legend { display:flex; gap:16px; flex-wrap:wrap; margin:6px 0 10px; color:#52514e; font-size:12px; }
.legend span { white-space:nowrap; }
.legend i { display:inline-block; width:14px; height:14px; border-radius:3px; vertical-align:-2px;
            margin-right:5px; border:1px solid rgba(11,11,11,.15); }
ol { margin:8px 0; } #findings li { margin:4px 0; }
.chip { display:inline-block; padding:1px 8px; border-radius:10px; font-size:11px; color:#0b0b0b;
        border:1px solid rgba(11,11,11,.10); margin-right:6px; }
.chip.l1{background:#9ec5f4;} .chip.l2{background:#fab219;}
.chip.l3{background:#ec835a;} .chip.l4{background:#d03b3b;color:#fff;}
.muted { color:#898781; }
details { margin:10px 0; } summary { cursor:pointer; padding:6px 0; }
"""


def _discrete_scale(colors: list[str]) -> list[list]:
    """Stepped plotly colorscale: integer value i -> solid colors[i]."""
    n = len(colors)
    cs = []
    for i, c in enumerate(colors):
        cs += [[i / n, c], [(i + 1) / n, c]]
    return cs


def _heatmap(z, x, y, text, colors, xlab, ylab, hovertemplate):
    import plotly.graph_objects as go
    n = len(colors)
    # colour key lives in an HTML legend next to the figure, not on the graph
    fig = go.Figure(go.Heatmap(
        z=z, x=x, y=y, text=text, hovertemplate=hovertemplate,
        zmin=0, zmax=n, colorscale=_discrete_scale(colors), xgap=1, ygap=1,
        showscale=False,
    ))
    fig.update_layout(template="plotly_white", font=dict(size=11),
                      margin=dict(l=10, r=10, t=10, b=10),
                      height=max(420, len(y) * 16 + 180),
                      xaxis_title=xlab, yaxis_title=ylab)
    fig.update_xaxes(showgrid=False, tickangle=-45, side="top", automargin=True)
    fig.update_yaxes(showgrid=False, autorange="reversed", automargin=True)
    return fig


def _uniq_labels(titles: list[str], width: int) -> list[str]:
    # index-prefix keeps labels unique (plotly merges rows/cols with identical labels)
    return [f"{i + 1}. {t[:width]}" for i, t in enumerate(titles)]


def _legend(colors: list[str], names: dict, absent: str | None = None) -> str:
    """HTML colour key shown beside a heat map (plotly colorbar is turned off)."""
    parts = [f'<span><i style="background:{c}"></i>{html.escape(names[i])}</span>'
             for i, c in enumerate(colors)]
    if absent:
        parts.append(f'<span><i style="background:#ececec"></i>{html.escape(absent)}</span>')
    return f'<div class="legend">{"".join(parts)}</div>'


def render_html(report: dict) -> str:
    files = report["files"]
    clusters = report["clusters"]
    cells = report["cells"]
    qvar = report["qvar"]
    ctitles = [c["title"] for c in clusters]
    esc = lambda s: html.escape(str(s))  # noqa: E731

    # severity heat map: rows = questionnaires, cols = questions.
    # hover text is just the status — the firm (y) and question (x) come from the
    # axes, so we don't duplicate the long title into all ~files*questions cells.
    z1, t1 = [], []
    for fn in files:
        zr, tr = [], []
        for lv in cells[fn]:
            zr.append(None if lv < 0 else lv)
            tr.append("" if lv < 0 else SEV_NAMES[lv])
        z1.append(zr); t1.append(tr)
    fig1 = _heatmap(z1, _uniq_labels(ctitles, 34), files, t1, SEV_COLORS,
                    "canonical question", "questionnaire",
                    "%{y}<br>%{x}<br><b>%{text}</b><extra></extra>")

    # wording heat map: rows = questions, cols = questionnaires
    wording = []
    for c in clusters:
        m = {}
        for v in c["variants"]:
            for f in v["files"]:
                m[f] = v["text"]
        wording.append(m)
    # transposed like the severity map: rows = questionnaires, cols = questions.
    # hover keeps status + this firm's actual wording (the value-add); firm (y)
    # and question (x) come from the axes.
    z2, t2 = [], []
    for fi, fn in enumerate(files):
        zr, tr = [], []
        for ci in range(len(clusters)):
            lv = qvar[ci][fi]
            if lv < 0:
                zr.append(None); tr.append("")
            else:
                zr.append(lv)
                tr.append(f"<b>{WORD_NAMES[lv]}</b><br>{esc(wording[ci].get(fn, '')[:140])}")
        z2.append(zr); t2.append(tr)
    fig2 = _heatmap(z2, _uniq_labels(ctitles, 34), files, t2, WORD_COLORS,
                    "canonical question", "questionnaire",
                    "%{y}<br>%{x}<br>%{text}<extra></extra>")

    cfg = {"displaylogo": False, "responsive": True}
    p1 = fig1.to_html(full_html=False, include_plotlyjs=True, config=cfg)
    p2 = fig2.to_html(full_html=False, include_plotlyjs=False, config=cfg)
    legend1 = _legend(SEV_COLORS, SEV_NAMES, "question not expected in this file")
    legend2 = _legend(WORD_COLORS, WORD_NAMES, "question not in this questionnaire")

    counts = Counter(g["level"] for g in report["findings"])
    tiles = (f'<div class="tile"><b>{len(files)}</b><span>questionnaires</span></div>'
             f'<div class="tile"><b>{len(clusters)}</b><span>canonical questions</span></div>'
             + "".join(f'<div class="tile"><b>{counts.get(l, 0)}</b><span>{SEV_NAMES[l]}</span></div>'
                       for l in (4, 3, 2, 1)))

    FCAP = 200
    items = [f'<li><span class="chip l{g["level"]}">{esc(g["kind"])}</span> '
             f'<b>{esc(g["file"])}</b> × {esc(ctitles[g["cluster"]][:60])} '
             f'— {esc(g["message"])}</li>' for g in report["findings"]]
    findings = f'<ol id="findings">{"".join(items[:FCAP])}</ol>'
    if len(items) > FCAP:
        findings += (f'<details><summary>Show all {len(items)} findings '
                     f'(most severe {FCAP} shown above)</summary>'
                     f'<ol start="{FCAP + 1}">{"".join(items[FCAP:])}</ol></details>')

    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>KYD questionnaire review</title>
<style>{PAGE_CSS}</style></head><body>
<h1>KYD questionnaire review</h1>
<p class="sub">Questions matched across files (ids not trusted). Hover any cell for the firm, question and status.</p>
<div class="tiles">{tiles}</div>
<h2>Findings by questionnaire &times; question</h2>
<p class="sub">Rows = questionnaires, columns = questions. Colour = worst issue found for that firm on that question.</p>
{legend1}
<div class="plot">{p1}</div>
<h2>Question wording vs. majority</h2>
<p class="sub">Rows = questionnaires, columns = questions. Green = same wording, yellow = formatting-only, red = substantially different.</p>
{legend2}
<div class="plot">{p2}</div>
<h2>All findings</h2>
{findings}
</body></html>"""


XLSX_FILLS = {OK: "DFF2DF", NOTE: "9EC5F4", WARN: "FAB219", MISSING: "EC835A", OUTLIER: "D03B3B"}


def render_xlsx(report: dict, path: Path) -> None:
    """Same content as the HTML, flat, for auditors who work in Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)

    def sheet(title, headers, rows, widths, fills=None):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        for ri, row in enumerate(rows, start=2):
            ws.append(row)
            for ci, fill in (fills(ri - 2) if fills else {}).items():
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", start_color=fill)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "B2"
        return ws

    clusters = report["clusters"]
    titles = [c["title"] for c in clusters]
    multi = report.get("multi_folder", False)
    folders = report.get("folders", {})
    lead = ["Folder", "File"] if multi else ["File"]
    off = len(lead) + 1

    ws = sheet("Heatmap", lead + titles,
               [([folders.get(fn, "")] if multi else []) + [fn] +
                [("n/a" if lv < 0 else LEVEL_NAMES[lv]) for lv in report["cells"][fn]]
                for fn in report["files"]],
               ([14] if multi else []) + [22] + [18] * len(titles),
               fills=lambda r: {ci + off: XLSX_FILLS[lv]
                                for ci, lv in enumerate(report["cells"][report["files"][r]]) if lv >= 0})
    for c in ws[1][len(lead):]:
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # --- wording heat map (firm x question), same layout as the HTML/Heatmap sheets:
    #     green = same as majority, yellow = formatting-only, red = substantially
    #     different, grey = question absent from that questionnaire ---
    WORD_FILLS = {0: "5CB85C", 1: "FAB219", 2: "D03B3B"}
    ABSENT_FILL = "ECECEC"
    qvar = report["qvar"]
    fnames = report["files"]
    wsw = sheet("Question wording", lead + titles,
                [([folders.get(fn, "")] if multi else []) + [fn] + [""] * len(clusters)
                 for fn in fnames],
                ([14] if multi else []) + [22] + [18] * len(titles),
                fills=lambda r: {ci + off: WORD_FILLS.get(qvar[ci][r], ABSENT_FILL)
                                 for ci in range(len(clusters))})
    for c in wsw[1][len(lead):]:
        c.alignment = Alignment(wrap_text=True, vertical="top")

    sheet("Findings", ["Severity", "Type", "File", "Question", "Detail"],
          [[LEVEL_NAMES[f["level"]], f["kind"], f["file"], titles[f["cluster"]], f["message"]]
           for f in report["findings"]],
          [16, 22, 22, 45, 90],
          fills=lambda r: {1: XLSX_FILLS[report["findings"][r]["level"]]})

    qrows = [(c, v) for c in clusters for v in c["variants"]]
    sheet("Questions", ["Question", "Coverage", "Specific?", "Variant type", "Wording (question + prompts)", "Files"],
          [[c["title"], f"{c['coverage']}/{len(report['files'])}", "yes" if c["specific"] else "",
            v["kind"], v["text"], ", ".join(sorted(v["files"]))] for c, v in qrows],
          [45, 10, 9, 14, 70, 40])

    arows = [(c, sl, e) for c in clusters for sl in c["slots"] for e in sl["entries"]]
    sheet("Answers", ["Question", "Sub-question", "File", "Id", "Selection", "Answer", "Flags"],
          [[c["title"], sl["label"], e["file"], c["ids"].get(e["file"], ""),
            e["selection"], e["answer"], ", ".join(e["flags"])] for c, sl, e in arows],
          [45, 40, 22, 6, 22, 70, 30])

    del wb["Sheet"]
    wb.save(path)


def main() -> None:
    import sys
    if hasattr(sys.stdout, "reconfigure"):  # Windows consoles default to a legacy codepage
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("jsons", nargs="*", help="parsed questionnaire JSON files")
    ap.add_argument("--dir", action="append", metavar="DIR",
                     help="review every *.json in this directory; repeat --dir to merge multiple source folders into one result")
    ap.add_argument("-o", "--out", default="output/kyd_review.html")
    ap.add_argument("--config", default="config.yaml", help="Azure config for embeddings")
    ap.add_argument("--embed-cache", default=".kyd_embed_cache.json",
                     help="on-disk embedding cache (reused across runs)")
    ap.add_argument("--no-embed", action="store_true",
                     help="skip embeddings; cluster by difflib text similarity only")
    args = ap.parse_args()

    dirs = args.dir or []
    multi_folder = len(dirs) > 1
    entries = [(p, None) for p in (Path(p) for p in args.jsons)]
    for d in dirs:
        dp = Path(d)
        stem = dp.resolve().name
        entries += [(f, stem if multi_folder else None) for f in _glob_json(dp)]
    if len(entries) < 2:
        ap.error("need at least 2 JSON files (pass paths or --dir, repeatable)")

    files = load_files(entries)
    vecs, threshold, cfg, config_error = None, None, {}, None
    try:
        cfg = load_config(args.config) or {}
    except Exception as e:
        config_error = e
    try:
        matching = question_matching_settings(cfg)
    except Exception as e:
        ap.error(str(e))
    if not args.no_embed:
        try:
            if config_error:
                raise config_error
            sigs = [norm(q_signature(q)) for f in files for q in f["questions"]]
            vecs = embed_signatures(sigs, cfg, Path(args.embed_cache))
            threshold = matching["embedding_threshold"]
            print(f"embeddings: {len(set(sigs))} unique signatures, semantic clustering "
                  f"({matching['mode']} profile, threshold {threshold:g})")
        except Exception as e:  # missing key/config/network -> lexical fallback
            print(f"embeddings unavailable ({e}); falling back to difflib text matching")
            vecs = None

    report = analyze(files, multi_folder, vecs, threshold, cfg.get("review", {}),
                     matching["slot_similarity"], matching["parent_presence"])
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(render_html(report), encoding="utf-8")
    xlsx = out.with_suffix(".xlsx")
    render_xlsx(report, xlsx)

    counts = Counter(g["level"] for g in report["findings"])
    print(f"{len(report['files'])} questionnaires, {len(report['clusters'])} canonical questions")
    for lv in (OUTLIER, MISSING, WARN, NOTE):
        print(f"  {LEVEL_NAMES[lv]:>18}: {counts.get(lv, 0)}")
    for g in report["findings"]:
        if g["level"] >= MISSING:
            print(f"  [{LEVEL_NAMES[g['level']]}] {g['file']} x {report['clusters'][g['cluster']]['title'][:50]}: {g['message']}")
    print(f"report -> {out} and {xlsx}")


if __name__ == "__main__":
    main()
