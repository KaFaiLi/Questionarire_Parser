"""Generate human-readable KYD questionnaires (xlsx, optional pdf) + ground-truth JSON.

These are SOURCE documents — the messy, merged-cell questionnaire layout the
LLM parser consumes — NOT the final parsed output. They reuse the same
question_ids across files but drift the question wording (mostly slight, a few
substantive) so downstream drift / consistency detection has something to chew
on.

Single source of truth: `build_questions(i)` returns the structured questions
(post-drift, post-fill). `render_xlsx` renders that to the questionnaire layout,
and the same structure is the GROUND TRUTH used to score LLM parsing accuracy.

Layout convention (mirrors Demo/demo_questionnaire*.xlsx):
    col B : selection (YES/NO) on option rows
    col L : question_id (Q1, Q2, ...)
    col N : section headers, question text, prompts, option labels, answers
Plus a distractor "reviewer note" cell off to the right the parser must ignore.

    python gen_kyd_questionnaires.py                 # 20 files, 11..30, xlsx + json
    python gen_kyd_questionnaires.py --pdf           # also export pdf (Excel COM)
    python gen_kyd_questionnaires.py --count 5 --start 11 --out some/dir

Deterministic: drift for file N is seeded by N.
"""
from __future__ import annotations

import argparse
import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# --- canonical questionnaire template -------------------------------------
# sections -> questions; each sub_question carries the final-format fields.
# Empty fields are filled in by _norm(). {placeholders} resolved per firm.
SECTIONS = [
    ("ORGANISATION, BUSINESS DESCRIPTION & EXPERIENCE", [
        ("Q1", "Firm Full Legal Name and Address", [
            {"answer": "{firm}"},
        ]),
        ("Q2", "Regulatory Authority", [
            {"prompt": "SFC, Monetary Authority of Singapore etc.", "answer": "{regulator}"},
        ]),
        ("Q3", "Relevant Authorised Activities", [
            {"prompt": "For HK, please specify license type", "answer": "Type 1 & Type 4"},
            {"prompt": "For Singapore, please specify license type", "answer": "N/A"},
        ]),
        ("Q4", "Economic data (If available: number of staff, AUM, ...)", [
            {"answer": "Approx no of employees: {staff}, AUM: {aum}"},
        ]),
    ]),
    ("PRODUCT APPROVAL PROCESS", [
        ("Q5", "Product approval process for structured products", [
            {"prompt": "a) Has your firm adopted a product approval process with respect to structured products?",
             "selection": "YES", "answer": "Yes"},
            {"prompt": "b) Does each structured product go through the same approval process?",
             "selection": "YES", "answer": "Yes, each structured product type will go through the same approval process."},
        ]),
        ("Q6", "What are the key features reviewed in the product approval process for structured products?", [
            {"answer": "Product risk classification, investor categorization, risk mapping to investor and product features."},
        ]),
    ]),
    ("DISTRIBUTION CHANNEL", [
        ("Q7", "In which capacity will your firm act as distributor?", [
            {"option_label": "Settling distributor", "selection": "YES"},
            {"prompt": "Who is the custodian?"},
            {"option_label": "Non-settling distributor", "selection": "NO"},
            {"prompt": "Who is the custodian?", "answer": "{custodian}"},
            {"prompt": "Is this PtP set up? i.e you act as principal rather than agent."},
        ]),
        ("Q8", "What marketing material you need?", [
            {"option_label": "SG Marketing material", "selection": "YES"},
            {"option_label": "Distributor marketing material prepared by your firm and not reviewed by SG", "selection": "{mkt2}"},
            {"option_label": "Distributor marketing material prepared by your firm and reviewed by SG", "selection": "{mkt3}"},
            {"option_label": "Other", "selection": "NO"},
        ]),
        ("Q9", "Does your firm directly distribute the products to the end investors?", [
            {"selection": "{direct}", "answer": "{direct_ans}"},
        ]),
    ]),
]

# substantive (meaning-changing) variants per question id
SUBSTANTIVE = {
    "Q1": [
        "Firm Full Legal Name and Registered Address",
        "Firm Full Legal Name, Address and Place of Incorporation",
    ],
    "Q4": ["Economic data (number of staff, AUM, and latest financial year revenue)"],
    "Q7": ["In which capacity will your firm act as distributor, and for which product lines?"],
    "Q9": ["Does your firm directly distribute the products to retail end investors?"],
}

SYNONYMS = [
    ("Firm", "Company"),
    ("Regulatory Authority", "Regulator"),
    ("structured products", "structured-product offerings"),
    ("marketing material", "marketing collateral"),
    ("distribute", "offer"),
]

# (firm name + address, regulator, staff, aum, custodian)
FIRMS = [
    ("Anchor Bay Securities (Hong Kong) Limited, 12 Connaught Road Central, HK", "SFC & HKMA", "320", "HK$ 55 billion", "Euroclear"),
    ("Brightwater Capital Pte Ltd, 8 Marina Boulevard, Singapore", "MAS", "210", "S$ 30 billion", "Clearstream"),
    ("Cobalt Crest Investments (HK) Limited, 5 Queen's Road Central, HK", "SFC", "180", "HK$ 22 billion", "HSBC"),
    ("Dunmore Asset Management Pte Ltd, 1 Raffles Place, Singapore", "MAS", "95", "S$ 12 billion", "Citibank"),
    ("Evergreen Trust Securities (Hong Kong) Limited, 18 Des Voeux Road, HK", "SFC & HKMA", "410", "HK$ 70 billion", "PERSHING LLC"),
    ("Fairhaven Brokerage Pte Ltd, 10 Collyer Quay, Singapore", "MAS", "60", "S$ 6 billion", "Standard Chartered"),
    ("Granite Peak Financial (HK) Limited, 22 Hennessy Road, HK", "SFC", "250", "HK$ 40 billion", "Euroclear"),
    ("Harborlight Wealth Pte Ltd, 3 Temasek Avenue, Singapore", "MAS", "140", "S$ 18 billion", "DBS"),
]

_FIELDS = ("option_label", "prompt", "selection", "answer")


def _norm(sub: dict, fill: dict) -> dict:
    return {f: str(sub.get(f, "")).format(**fill) for f in _FIELDS}


def _slight(text: str, rng: random.Random) -> str:
    c = rng.randrange(4)
    if c == 0:
        return text.replace(" ", "  ", 1)      # doubled space (formatting drift)
    if c == 1:
        return text.upper()                    # case drift
    if c == 2:
        for a, b in SYNONYMS:
            if a in text:
                return text.replace(a, b, 1)   # synonym swap
    return text + " (if applicable)"           # appended qualifier


def _drift(qid: str, text: str, rng: random.Random) -> str:
    r = rng.random()
    if r < 0.70:
        return text
    if r < 0.92:
        return _slight(text, rng)
    subs = SUBSTANTIVE.get(qid)
    return rng.choice(subs) if subs else _slight(text, rng)


def build_questions(file_index: int) -> dict:
    """Ground-truth structured questionnaire for a given file index."""
    rng = random.Random(file_index)
    firm, regulator, staff, aum, custodian = FIRMS[file_index % len(FIRMS)]
    direct = rng.choice(["YES", "YES", "YES", "NO"])  # mostly YES, occasional outlier
    fill = {
        "firm": firm, "regulator": regulator, "staff": staff, "aum": aum,
        "custodian": custodian,
        "mkt2": rng.choice(["YES", "NO"]),
        "mkt3": rng.choice(["YES", "NO"]),
        "direct": direct,
        "direct_ans": "Yes" if direct == "YES" else "No",
    }
    questions = []
    for _section, qs in SECTIONS:
        for qid, text, subs in qs:
            questions.append({
                "question_id": qid,
                "question": _drift(qid, text, rng).format(**fill),
                "sub_questions": [_norm(s, fill) for s in subs],
            })
    return {"file_name": f"{file_index:02d}_questionnaire", "questions": questions}


def render_xlsx(qdata: dict) -> Workbook:
    """Render structured questions into the questionnaire layout."""
    by_id = {q["question_id"]: q for q in qdata["questions"]}
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire HK"
    bold = Font(bold=True)
    row = 2
    noise_done = False
    for section, qs in SECTIONS:
        ws.cell(row=row, column=14, value=section).font = bold   # col N
        row += 2
        for qid, _text, _subs in qs:
            q = by_id[qid]
            ws.cell(row=row, column=12, value=qid)               # col L
            ws.cell(row=row, column=14, value=q["question"]).font = bold
            row += 1
            for sub in q["sub_questions"]:
                if sub["option_label"]:
                    if sub["selection"]:
                        ws.cell(row=row, column=2, value=sub["selection"])   # col B
                    ws.cell(row=row, column=14, value=sub["option_label"])
                    row += 1
                if sub["prompt"]:
                    ws.cell(row=row, column=14, value=sub["prompt"])
                    row += 1
                if sub["answer"]:
                    # selection w/o option_label sits next to the answer
                    if sub["selection"] and not sub["option_label"]:
                        ws.cell(row=row, column=2, value=sub["selection"])
                    ws.cell(row=row, column=14, value=sub["answer"])
                    row += 1
            row += 1  # blank spacer between questions
            if not noise_done and row > 9:
                ws.cell(row=row - 1, column=16,   # col P: off to the side but keeps print width sane
                        value="External reviewer note: please verify Q1 address.")
                noise_done = True
    return wb


def to_pdf(xlsx_path: Path) -> Path | None:
    """Export xlsx -> pdf via Excel COM (Windows + Excel only). None on failure."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        return None
    pdf_path = xlsx_path.with_name(f"{xlsx_path.stem}__Questionnaire HK.pdf")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        ws = wb.Worksheets(1)
        ps = ws.PageSetup
        ps.Zoom = False            # required before FitToPages takes effect
        ps.FitToPagesWide = 1      # keep the whole form on a single page width
        ps.FitToPagesTall = False  # allow it to flow down multiple pages if needed
        ps.Orientation = 1         # 1 = portrait
        wb.ExportAsFixedFormat(0, str(pdf_path.resolve()))  # 0 = xlTypePDF
        wb.Close(False)
        return pdf_path
    finally:
        excel.Quit()


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--count", type=int, default=20)
    ap.add_argument("--start", type=int, default=11)
    ap.add_argument("--out", default="Demo/generated_questionnaires")
    ap.add_argument("--pdf", action="store_true", help="also export pdf via Excel COM")
    ap.add_argument("--no-json", action="store_true", help="skip ground-truth json")
    args = ap.parse_args()

    out = Path(args.out)
    gt = out / "ground_truth"
    out.mkdir(parents=True, exist_ok=True)
    if not args.no_json:
        gt.mkdir(parents=True, exist_ok=True)

    for i in range(args.start, args.start + args.count):
        qdata = build_questions(i)
        xlsx = out / f"{i:02d}_questionnaire.xlsx"
        render_xlsx(qdata).save(xlsx)
        line = f"wrote {xlsx.name}"
        if not args.no_json:
            (gt / f"{i:02d}_questionnaire.json").write_text(
                json.dumps(qdata, indent=2, ensure_ascii=False), encoding="utf-8")
            line += " + ground_truth json"
        if args.pdf:
            pdf = to_pdf(xlsx)
            line += f" + {pdf.name}" if pdf else "  (pdf skipped: Excel COM unavailable)"
        print(line)


def _selfcheck() -> None:
    """ponytail: minimal runnable check — structure, determinism, drift."""
    q11 = build_questions(11)
    ids = [q["question_id"] for q in q11["questions"]]
    assert ids == ["Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9"], ids
    for q in q11["questions"]:
        for s in q["sub_questions"]:
            assert set(s) == set(_FIELDS), s
    import re as _re
    assert not _re.search(r"\{[a-z_]+\}", json.dumps(q11)), "unresolved placeholder"
    assert build_questions(11) == build_questions(11), "non-deterministic"
    assert build_questions(11) != build_questions(12), "no drift between files"
    # xlsx round-trips the ids
    from io import BytesIO
    from openpyxl import load_workbook
    buf = BytesIO(); render_xlsx(q11).save(buf); buf.seek(0)
    ws = load_workbook(buf).active
    seen = {c.value for r in ws.iter_rows() for c in r if c.value in set(ids)}
    assert seen == set(ids), seen
    print("selfcheck OK")


if __name__ == "__main__":
    import sys
    if "--selfcheck" in sys.argv:
        _selfcheck()
    else:
        main()
