"""Planted scenarios in Demo/kyd_examples must all be detected (see its README)."""
from pathlib import Path

import kyd_review as kr

# difflib path: deterministic, no Azure needed (embeddings covered separately)
REPORT = kr.analyze(kr.load_files([(p, None) for p in sorted(Path("Demo/kyd_examples").glob("*.json"))]))


def _findings(kind=None, file=None):
    return [f for f in REPORT["findings"]
            if (kind is None or f["kind"] == kind) and (file is None or file in f["file"])]


def _cluster_of(f):
    return REPORT["clusters"][f["cluster"]]["title"]


def test_ids_not_trusted_capacity_question_merges_across_hk_sg():
    c = next(c for c in REPORT["clusters"] if "capacity" in c["title"])
    assert c["coverage"] == 10 and {"Q10", "Q12"} <= set(c["ids"].values())


def test_tolerant_question_matching_profile():
    tolerant = kr.question_matching_settings({"question_matching": {"mode": "tolerant"}})
    assert tolerant["mode"] == "tolerant"
    assert tolerant["embedding_threshold"] == 0.78
    assert tolerant["slot_similarity"] == 0.60
    assert tolerant["parent_presence"] == {
        "allow_top_level_evidence": True,
        "top_level_similarity": 0.78,
        "allow_subquestion_evidence": True,
        "slot_similarity": 0.78,
        "minimum_distinctive_slot_matches": 1,
        "uniqueness_margin": 0.05,
        "ignore_subquestion_wording_drift": True,
    }

    # Standard mode preserves the legacy drift threshold unless explicitly
    # overridden under question_matching.
    standard = kr.question_matching_settings({"drift": {"cluster_threshold": 0.88}})
    assert standard["embedding_threshold"] == 0.88
    assert standard["slot_similarity"] == kr.SIM_SLOT
    assert not standard["parent_presence"]["allow_subquestion_evidence"]


def test_tolerant_slot_alignment_keeps_enumerated_siblings_separate():
    def question(label, prompt):
        return {"question_id": "Q1", "question": "", "sub_questions": [
            {"option_label": label, "prompt": prompt, "selection": "", "answer": ""}]}

    c = {"members": {
        "one": question("", "Provide details of anti money laundering policy"),
        "two": question("", "Anti money laundering policy details"),
    }}
    assert len(kr.align_slots(c, 0.70)) == 2
    assert len(kr.align_slots(c, 0.60)) == 1

    siblings = {"members": {
        "one": question("a)", "Provide details of anti money laundering policy"),
        "two": question("b)", "Anti money laundering policy details"),
    }}
    assert len(kr.align_slots(siblings, 0.60)) == 2


def test_tolerant_parent_presence_ignores_partial_child_parsing_and_drift():
    def question(prompts):
        return {"question_id": "Q1", "question": "AML controls", "sub_questions": [
            {"option_label": label, "prompt": prompt, "selection": "", "answer": ""}
            for label, prompt in prompts]}

    full = [("a)", "Entity legal name"), ("b)", "Details of anti money laundering controls")]
    files = [
        {"name": "one", "questions": [question(full)]},
        # The heading is intact but the child parse is incomplete/different.
        # Top-level evidence must prove the parent is present.
        {"name": "two", "questions": [question([("a)", "Parser extraction fragment")])]},
        {"name": "three", "questions": [question(full)]},
    ]
    presence = kr.question_matching_settings({"question_matching": {"mode": "tolerant"}})["parent_presence"]
    report = kr.analyze(files, threshold=0.78, parent_presence=presence)

    assert len(report["clusters"]) == 1
    assert report["clusters"][0]["coverage"] == 3
    assert not [f for f in report["findings"] if f["kind"] in {"question missing", "wording variant"}]
    # The parent exists in every file; the absent b) child is merely a partial
    # parse and is not a missing-question finding.
    slots = report["clusters"][0]["slots"]
    assert any({e["file"] for e in s["entries"]} == {"one", "three"} for s in slots)


def test_tolerant_parent_presence_falls_back_to_distinctive_child_when_heading_blank():
    def question(prompts):
        return {"question_id": "Q1", "question": "", "sub_questions": [
            {"option_label": label, "prompt": prompt, "selection": "", "answer": ""}
            for label, prompt in prompts]}

    full = [("a)", "Entity legal name"), ("b)", "Details of anti money laundering controls")]
    files = [
        {"name": "one", "questions": [question(full)]},
        {"name": "two", "questions": [question([("a)", "Entity legal name")])]},
        {"name": "three", "questions": [question(full)]},
    ]
    presence = kr.question_matching_settings({"question_matching": {"mode": "tolerant"}})["parent_presence"]
    report = kr.analyze(files, threshold=0.78, parent_presence=presence)

    assert len(report["clusters"]) == 1
    assert report["clusters"][0]["coverage"] == 3


def test_parent_presence_keeps_enumerated_siblings_separate():
    def question(prompts):
        return {"question_id": "Q1", "question": "", "sub_questions": [
            {"option_label": label, "prompt": prompt, "selection": "", "answer": ""}
            for label, prompt in prompts]}

    full = [("a)", "Entity legal name"), ("b)", "Details of anti money laundering controls")]
    files = [
        {"name": "one", "questions": [question(full)]},
        # The same text under b) is not proof of the a) child/parent mapping.
        {"name": "two", "questions": [question([("b)", "Entity legal name")])]},
        {"name": "three", "questions": [question(full)]},
    ]
    presence = kr.question_matching_settings({"question_matching": {"mode": "tolerant"}})["parent_presence"]
    report = kr.analyze(files, threshold=0.78, parent_presence=presence)

    assert sorted(c["coverage"] for c in report["clusters"]) == [1, 2]


def test_parent_presence_rejects_generic_child_prompts():
    def question(prompts):
        return {"question_id": "Q1", "question": "", "sub_questions": [
            {"option_label": label, "prompt": prompt, "selection": "", "answer": ""}
            for label, prompt in prompts]}

    files = [
        {"name": "one", "questions": [question([("a)", "Please provide details"),
                                                     ("b)", "Describe governance arrangements")])]},
        {"name": "two", "questions": [question([("a)", "Please provide details")])]},
        {"name": "three", "questions": [question([("a)", "Please provide details"),
                                                       ("b)", "Describe governance arrangements")])]},
    ]
    presence = kr.question_matching_settings({"question_matching": {"mode": "tolerant"}})["parent_presence"]
    report = kr.analyze(files, threshold=0.78, parent_presence=presence)

    assert sorted(c["coverage"] for c in report["clusters"]) == [1, 2]


def test_questionnaire_specific_mas_question():
    fs = _findings("questionnaire-specific")
    assert len(fs) == 1 and fs[0]["file"] == "02_pacific_trust_sg"


def test_focused_wording_policy_and_formatting_context():
    # The planted Registered Address variation belongs to a questionnaire-only
    # question, so focused review keeps it in the wording grid but does not
    # promote it to a finding.
    assert not any("Registered Address" in f["message"] for f in _findings("wording variant"))
    assert _findings("formatting variant", "06_summit")   # extra spaces in Q1
    assert _findings("formatting variant", "09_apex")     # REGULATORY  AUTHORITY
    for f in _findings("wording variant"):
        c = REPORT["clusters"][f["cluster"]]
        assert c["coverage"] / len(REPORT["files"]) >= kr.WORDING_MIN_COVERAGE
        variant = next(v for v in c["variants"] if f["file"] in v["files"])
        assert len(variant["files"]) / len(REPORT["files"]) <= kr.WORDING_VARIANT_MAX_SHARE


def test_missing_required_answers():
    assert any("conomic data" in _cluster_of(f) for f in _findings("missing answer", "04_crescent"))
    assert any("Marketing" in _cluster_of(f) for f in _findings("missing answer", "06_summit"))


def test_conditional_blanks_not_flagged():
    assert not any("custodian" in _cluster_of(f) for f in _findings("missing answer"))


def test_answer_outliers():
    assert _findings("selection outlier", "07_harbourview")   # 9x YES / 1x NO
    assert _findings("free-text outlier", "05_northwind")     # divergent marketing answer
    assert _findings("numeric outlier", "03_meridian")        # 45 issuers vs 8-12


def test_declared_na_flagged():
    fs = _findings("declared N/A")
    assert len(fs) == 1 and fs[0]["file"] == "08_silkroad_sg"
    assert fs[0]["level"] == kr.WARN
    assert "conomic data" in _cluster_of(fs[0])


def test_yes_no_synonym_not_an_outlier():
    answer_kinds = {"selection outlier", "free-text outlier", "numeric outlier",
                    "selection/answer mismatch", "declared N/A", "missing answer"}
    assert not [f for f in _findings(file="10_lighthouse") if f["kind"] in answer_kinds]
    # and canonicalization must not hide the real 9-vs-1 NO outlier
    assert _findings("free-text outlier", "07_harbourview")


def test_heatmap_matrix_shape():
    assert len(REPORT["cells"]) == 10
    assert all(len(v) == len(REPORT["clusters"]) for v in REPORT["cells"].values())


def test_qvar_wording_grid():
    # question x firm wording-divergence grid: codes 0 same / 1 formatting / 2 substantive / -1 absent
    qv = REPORT["qvar"]
    assert len(qv) == len(REPORT["clusters"])
    assert all(len(row) == len(REPORT["files"]) for row in qv)
    codes = {c for row in qv for c in row}
    assert codes <= {-1, 0, 1, 2}
    # The grid preserves substantive differences as context; focused policy can
    # deliberately suppress them from the findings list when coverage is low.
    assert any(code == 2 for row in qv for code in row)


def test_xlsx_export(tmp_path):
    from openpyxl import load_workbook
    p = tmp_path / "r.xlsx"
    kr.render_xlsx(REPORT, p)
    wb = load_workbook(p)
    assert [ws.title for ws in wb] == ["Heatmap", "Question wording", "Findings", "Questions", "Answers"]
    assert wb["Findings"].max_row == len(REPORT["findings"]) + 1
    assert wb["Heatmap"].max_column == len(REPORT["clusters"]) + 1
    # wording grid: one row per firm, one column per question (+ the File column)
    wsw = wb["Question wording"]
    assert wsw.max_row == len(REPORT["files"]) + 1
    assert wsw.max_column == len(REPORT["clusters"]) + 1
    # a substantive cell (qvar==2) must be filled red
    for ci, row in enumerate(REPORT["qvar"]):
        for fi, code in enumerate(row):
            if code == 2:
                assert wsw.cell(row=fi + 2, column=ci + 2).fill.start_color.rgb.endswith("D03B3B")
                return
