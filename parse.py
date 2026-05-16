"""
Excel Questionnaire Parser
==========================
Parses questionnaire .xlsx files with merged cells, coloured fills, and dropdowns.

Key concept
-----------
A grey- or yellow-filled cell is the *answer box* (even if empty). It marks the
boundary between sub-questions. Everything else (plain text, green-highlighted
text, italic hints) belongs to the current sub-question's prompt.

So one Q-block looks like:
    [prompt text...] → [grey/yellow answer slot] → [prompt text...] → [grey slot] → ...
and each prompt+slot pair becomes one SubQuestion.

Usage:
    python parse_questionnaire.py <file_or_dir> [output_dir] [--debug]

Outputs:
    Single file input:
        - <name>.json : full structured data
        - <name>.xlsx : one row per sub-question (flattened)
    Folder input:
        - <stem>.json per input file (full structured data)
        - <folder>_<YYYY-MM-DD>.xlsx : combined flattened rows from all files

Requires: openpyxl  (pip install openpyxl)
"""

from __future__ import annotations
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re, json, sys, argparse
from datetime import date
from pathlib import Path
from dataclasses import dataclass, field, asdict

# ───── Configuration ────────────────────────────────────────────────────────
CONTENT_START_COL = column_index_from_string("B")     # 2
CONTENT_END_COL   = column_index_from_string("AO")    # 41
SIDE_ANS_END_COL  = column_index_from_string("K")     # side-answer dropdowns live left of Q-number col

Q_PATTERN = re.compile(r"^Q(\d+)\s*$")

# Fill-colour classifier (tune RGB ranges if real files use different shades)
COLOUR_BUCKETS = {
    "grey":   lambda r, g, b: abs(r - g) < 20 and abs(g - b) < 20 and 150 <= r <= 245,
    "green":  lambda r, g, b: g > r + 10 and g > b + 10 and g >= 150,
    "yellow": lambda r, g, b: r >= 200 and g >= 180 and b < 170 and abs(r - g) < 60,
}
ANSWER_FILLS = {"grey", "yellow"}   # fills that mark an answer slot

# ───── Data classes ─────────────────────────────────────────────────────────
@dataclass
class ContentBlock:
    text: str
    cell: str
    italic: bool = False
    bold: bool = False
    fill: str = "none"     # grey / green / yellow / none
    role: str = ""         # prompt / hint / answer

@dataclass
class SubQuestion:
    prompt: str = ""                     # all non-answer text leading up to the answer slot
    hint: str = ""                       # convenience: italic text inside the prompt
    answer: str = ""                     # text in the grey/yellow answer slot (may be empty)
    answer_cell: str = ""
    side_answer: str = ""                # dropdown value (e.g. YES / NO)
    side_answer_options: str = ""
    side_answer_cell: str = ""
    blocks: list = field(default_factory=list)   # raw breakdown for debugging / edge cases

@dataclass
class QAItem:
    section: str = ""
    question_id: str = ""
    sub_questions: list = field(default_factory=list)   # list[SubQuestion]
    source_rows: list = field(default_factory=list)

@dataclass
class Remark:
    cell: str
    value: str

# ───── Helpers ──────────────────────────────────────────────────────────────
def build_merged_lookup(ws):
    """(row, col) -> (anchor_row, anchor_col) for every cell inside a merged range."""
    lookup = {}
    for mr in ws.merged_cells.ranges:
        anchor = (mr.min_row, mr.min_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                lookup[(r, c)] = anchor
    return lookup

def build_dropdown_map(ws):
    """(row, col) -> options string for every cell with a list-style data validation."""
    dd = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        opts = (dv.formula1 or "").strip('"')
        for rng in dv.sqref.ranges:
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    dd[(r, c)] = opts
    return dd

def effective_value(ws, r, c, merged_lookup):
    if (r, c) in merged_lookup:
        ar, ac = merged_lookup[(r, c)]
        return ws.cell(ar, ac).value
    return ws.cell(r, c).value

def classify_fill(cell) -> str:
    fg = getattr(cell.fill, "fgColor", None)
    if not fg or fg.type != "rgb" or not fg.rgb:
        return "none"
    rgb = fg.rgb[-6:]
    try:
        r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return "none"
    for name, fn in COLOUR_BUCKETS.items():
        if fn(r, g, b):
            return name
    return "none"

def is_section_header(cell, value: str) -> bool:
    if not (cell.font and cell.font.bold):
        return False
    size = cell.font.sz or 11
    return size >= 12 and value.isupper() and len(value) > 4

def append_line(existing: str, new: str) -> str:
    return (existing + "\n" + new).strip() if existing else new

# ───── Core parser ──────────────────────────────────────────────────────────
def parse_workbook(filepath: Path, debug: bool = False):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_out = {}

    for ws in wb.worksheets:
        if debug:
            print(f"\n=== Sheet: {ws.title} ===")

        merged    = build_merged_lookup(ws)
        dropdowns = build_dropdown_map(ws)
        items: list[QAItem]  = []
        remarks: list[Remark] = []
        section = ""
        current_q:  QAItem | None       = None
        current_sq: SubQuestion | None  = None
        seen_anchors: set = set()

        def is_meaningful(sq: SubQuestion) -> bool:
            return bool(sq.prompt or sq.answer or sq.side_answer or sq.answer_cell)

        def flush_subq():
            nonlocal current_sq
            if current_q is not None and current_sq is not None and is_meaningful(current_sq):
                current_q.sub_questions.append(current_sq)
            current_sq = SubQuestion() if current_q is not None else None

        def flush_question():
            nonlocal current_q, current_sq
            if current_q is not None:
                if current_sq is not None and is_meaningful(current_sq):
                    current_q.sub_questions.append(current_sq)
                items.append(current_q)
            current_q = None
            current_sq = None

        for row in range(1, ws.max_row + 1):
            # Collect cells in the content range. Include cells with text OR with an
            # "answer fill" (grey/yellow) even if empty — those are empty answer slots.
            row_cells = []
            for col in range(CONTENT_START_COL, CONTENT_END_COL + 1):
                anchor = merged.get((row, col), (row, col))
                if anchor in seen_anchors:
                    continue
                cell = ws.cell(anchor[0], anchor[1])
                val  = cell.value
                fill = classify_fill(cell)
                text = "" if val is None else str(val).strip()
                if not text and fill not in ANSWER_FILLS:
                    continue
                seen_anchors.add(anchor)
                row_cells.append((col, text, cell, fill))

            row_kind = "empty"

            if row_cells:
                first_col, first_text, first_cell, first_fill = row_cells[0]

                # ── Section header ────────────────────────────────────────
                if first_text and is_section_header(first_cell, first_text):
                    flush_question()
                    section = first_text
                    row_kind = f"SECTION → {section!r}"

                else:
                    # ── New question? ─────────────────────────────────────
                    q_hit = next(((c, t, cell)
                                  for c, t, cell, _ in row_cells
                                  if t and Q_PATTERN.match(t)), None)

                    if q_hit:
                        q_col, q_text, _ = q_hit
                        flush_question()
                        m = Q_PATTERN.match(q_text)
                        current_q = QAItem(section=section,
                                           question_id=m.group(0),
                                           source_rows=[row])
                        current_sq = SubQuestion()

                        # Anything to the right of the Q-number cell starts the
                        # first sub-question's prompt.
                        for c, t, cell, fill in row_cells:
                            if c <= q_col or not t:
                                continue
                            italic = bool(cell.font and cell.font.italic)
                            bold   = bool(cell.font and cell.font.bold)
                            block = ContentBlock(text=t, cell=cell.coordinate,
                                                 italic=italic, bold=bold,
                                                 fill=fill, role="prompt")
                            current_sq.prompt = append_line(current_sq.prompt, t)
                            if italic:
                                current_sq.hint = append_line(current_sq.hint, t)
                            current_sq.blocks.append(block)
                        row_kind = (f"QUESTION {current_q.question_id}: "
                                    f"{current_sq.prompt[:70]!r}")

                    elif current_q is not None:
                        # ── Content for current question ──────────────────
                        current_q.source_rows.append(row)
                        if current_sq is None:
                            current_sq = SubQuestion()
                        summaries = []
                        for c, t, cell, fill in row_cells:
                            italic = bool(cell.font and cell.font.italic)
                            bold   = bool(cell.font and cell.font.bold)
                            block = ContentBlock(text=t, cell=cell.coordinate,
                                                 italic=italic, bold=bold, fill=fill)

                            if fill in ANSWER_FILLS:
                                # ── This is an answer slot (even if empty) ──
                                block.role = "answer"
                                current_sq.answer = append_line(current_sq.answer, t) if t else current_sq.answer
                                current_sq.answer_cell = cell.coordinate
                                current_sq.blocks.append(block)
                                summaries.append(f"{cell.coordinate}[ANSWER,{fill}]={t[:30]!r}")
                                # Close this sub-question; start a fresh one
                                flush_subq()
                            elif t:
                                # Plain text or green emphasis — part of the prompt
                                block.role = "hint" if italic else "prompt"
                                current_sq.prompt = append_line(current_sq.prompt, t)
                                if italic:
                                    current_sq.hint = append_line(current_sq.hint, t)
                                current_sq.blocks.append(block)
                                summaries.append(
                                    f"{cell.coordinate}[{block.role},{fill}"
                                    f"{',it' if italic else ''}]"
                                )
                        if summaries:
                            row_kind = "CONTENT: " + " | ".join(summaries)

                # ── Side-answer dropdown on this row ──────────────────────
                if current_q is not None:
                    for col in range(CONTENT_START_COL, SIDE_ANS_END_COL + 1):
                        if (row, col) not in dropdowns:
                            continue
                        val = effective_value(ws, row, col, merged)
                        if not val:
                            continue
                        if current_sq is None:
                            current_sq = SubQuestion()
                        current_sq.side_answer        = str(val).strip()
                        current_sq.side_answer_options = dropdowns[(row, col)]
                        current_sq.side_answer_cell   = f"{get_column_letter(col)}{row}"
                        row_kind += (f"  ⊕ SIDE[{current_sq.side_answer_cell}="
                                     f"{current_sq.side_answer!r}]")
                        break

            # ── Remarks: anything OUTSIDE B:AO ────────────────────────────
            outside_cols = list(range(1, CONTENT_START_COL)) + \
                           list(range(CONTENT_END_COL + 1, ws.max_column + 1))
            for col in outside_cols:
                anchor = merged.get((row, col), (row, col))
                if anchor in seen_anchors:
                    continue
                val = effective_value(ws, row, col, merged)
                if val is None:
                    continue
                text = str(val).strip()
                if text:
                    seen_anchors.add(anchor)
                    cell_ref = f"{get_column_letter(col)}{row}"
                    remarks.append(Remark(cell=cell_ref, value=text))
                    if debug:
                        print(f"  row {row:>4}  REMARK {cell_ref}: {text[:60]!r}")

            if debug and row_kind != "empty":
                print(f"  row {row:>4}  {row_kind}")

        flush_question()
        sheets_out[ws.title] = {"items": items, "remarks": remarks}

    return sheets_out

# ───── Output writers ───────────────────────────────────────────────────────
def write_json(sheets_out, path: Path):
    data = {
        sheet: {
            "questions": [asdict(it) for it in payload["items"]],
            "remarks":   [asdict(r)  for r  in payload["remarks"]],
        }
        for sheet, payload in sheets_out.items()
    }
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

XLSX_FIELDS = ["file_name", "sheet", "section", "question_id", "sub_idx",
               "prompt", "hint", "answer", "answer_cell",
               "side_answer", "side_answer_options", "side_answer_cell"]

def _iter_rows(file_name: str, sheets_out):
    for sheet, payload in sheets_out.items():
        for it in payload["items"]:
            for idx, sq in enumerate(it.sub_questions):
                yield [
                    file_name,
                    sheet,
                    it.section,
                    it.question_id,
                    idx,
                    sq.prompt,
                    sq.hint,
                    sq.answer,
                    sq.answer_cell,
                    sq.side_answer,
                    sq.side_answer_options,
                    sq.side_answer_cell,
                ]

def write_xlsx(entries: list[tuple[str, dict]], path: Path):
    """Flatten: one row per sub-question. entries is [(file_name, sheets_out), ...]."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "questions"
    ws.append(XLSX_FIELDS)
    for file_name, sheets_out in entries:
        for row in _iter_rows(file_name, sheets_out):
            ws.append(row)
    wb.save(path)

# ───── CLI ──────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Parse Excel questionnaires.")
    ap.add_argument("target", help="Path to .xlsx file OR directory of .xlsx files")
    ap.add_argument("output_dir", nargs="?", default="./output",
                    help="Output directory (default: ./output)")
    ap.add_argument("--debug", action="store_true",
                    help="Print row-by-row classification while parsing")
    args = ap.parse_args()

    target  = Path(args.target)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    is_folder = target.is_dir()
    files = sorted(target.glob("*.xlsx")) if is_folder else [target]
    if not files:
        print(f"No .xlsx files found in {target}"); sys.exit(1)

    entries: list[tuple[str, dict]] = []
    for fp in files:
        print(f"→ Parsing {fp.name}")
        result = parse_workbook(fp, debug=args.debug)
        write_json(result, out_dir / f"{fp.stem}.json")
        entries.append((fp.stem, result))

    if is_folder:
        xlsx_path = out_dir / f"{target.name}_{date.today().isoformat()}.xlsx"
    else:
        xlsx_path = out_dir / f"{files[0].stem}.xlsx"
    write_xlsx(entries, xlsx_path)
    print(f"\n✓ Done. Output in {out_dir.resolve()}")

if __name__ == "__main__":
    main()